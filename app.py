import os, io, csv, secrets
from datetime import datetime
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file, flash, abort, session
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                  TableStyle, HRFlowable, KeepTogether, Image)
from reportlab.graphics.shapes import Drawing, Rect, String, Line
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics import renderPDF

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

try:
    from flask_limiter import Limiter
    from flask_limiter.util import get_remote_address
    _limiter_available = True
except ImportError:
    _limiter_available = False

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', '')
if not app.secret_key:
    raise RuntimeError('SECRET_KEY non impostata. Aggiungila al file .env')

# ── Filtri Jinja2 per formattazione numeri IT (punto migliaia, virgola decimali) ──
def _it(value, dec=2):
    """Formatta un numero in stile italiano: 1.234,56"""
    if value is None:
        return '—'
    try:
        s = f'{float(value):,.{dec}f}'          # es. '1,234.56'
        return s.replace(',', 'X').replace('.', ',').replace('X', '.')  # → '1.234,56'
    except (ValueError, TypeError):
        return str(value)

app.jinja_env.filters['n2']  = lambda v: _it(v, 2)   # 2 decimali (default €)
app.jinja_env.filters['n4']  = lambda v: _it(v, 4)   # 4 decimali (spread/consumo)
app.jinja_env.filters['n1']  = lambda v: _it(v, 1)   # 1 decimale (percentuali)
app.jinja_env.filters['n0']  = lambda v: _it(v, 0)   # interi

login_manager = LoginManager(app)
login_manager.login_view = 'login'

ADMIN_EMAILS = {e.strip().lower() for e in os.environ.get('ADMIN_EMAILS', 'fvedani23@gmail.com').split(',')}
REGISTRATION_OPEN = os.environ.get('REGISTRATION_OPEN', 'false').lower() == 'true'
MIN_PASSWORD_LEN  = 8

# ── Flask-Limiter (opzionale — installare flask-limiter) ────────────────────────
if _limiter_available:
    limiter = Limiter(get_remote_address, app=app, default_limits=[],
                      storage_uri='memory://')
else:
    limiter = None

# ── CSRF protection (sessione) ──────────────────────────────────────────────────
def _csrf_token():
    if '_csrf' not in session:
        session['_csrf'] = secrets.token_hex(32)
    return session['_csrf']

app.jinja_env.globals['csrf_token'] = _csrf_token

@app.before_request
def _csrf_check():
    if request.method not in ('POST', 'PUT', 'PATCH', 'DELETE'):
        return
    if request.is_json:
        return
    token = request.form.get('_csrf') or request.headers.get('X-CSRF-Token', '')
    if not token or not secrets.compare_digest(token, session.get('_csrf', '')):
        abort(403)

CONSUMI_DEFAULT = {
    ('DOMESTICO', 'LUCE'): 2.7,
    ('DOMESTICO', 'GAS'):  800,
    ('ALTRI USI', 'LUCE'): 15,
    ('ALTRI USI', 'GAS'):  5000,
    ('CONDOMINI', 'LUCE'): 7,
    ('CONDOMINI', 'GAS'):  20000,
}

import re
import psycopg2
import psycopg2.extras


class RowWrapper(dict):
    """Dict compatibile con sqlite3.Row: supporta row['col'] e row[0]."""
    def __getitem__(self, key):
        if isinstance(key, int):
            return list(self.values())[key]
        return super().__getitem__(key)


class CursorWrapper:
    def __init__(self, cur):
        self._cur = cur
        self.lastrowid = None

    def fetchone(self):
        row = self._cur.fetchone()
        return RowWrapper(row) if row is not None else None

    def fetchall(self):
        return [RowWrapper(r) for r in (self._cur.fetchall() or [])]

    def __iter__(self):
        return iter(self.fetchall())


class DbWrapper:
    """Wrapper psycopg2 con interfaccia compatibile sqlite3."""

    def __init__(self, conn):
        self._conn = conn
        self.lastrowid = None

    @staticmethod
    def _translate(query):
        # ? -> %s (parametri)
        q = query.replace('?', '%s')
        # "VALORE_CAPS" -> 'VALORE_CAPS' (PostgreSQL vuole apici singoli per stringhe)
        q = re.sub(r'"([A-Z][A-Z0-9_ ]*)"', r"'\1'", q)
        return q

    def execute(self, query, params=None):
        q = self._translate(query)
        cur = self._conn.cursor()
        cur.execute(q, params if params is not None else [])
        wrapper = CursorWrapper(cur)
        if q.strip().upper().startswith('INSERT'):
            try:
                lv = self._conn.cursor()
                lv.execute('SELECT lastval()')
                row = lv.fetchone()
                lid = row['lastval'] if row else None
                wrapper.lastrowid = lid
                self.lastrowid = lid
            except Exception:
                pass
        return wrapper

    def executescript(self, sql):
        for stmt in sql.split(';'):
            stmt = stmt.strip()
            if not stmt:
                continue
            try:
                cur = self._conn.cursor()
                cur.execute(stmt)
                self._conn.commit()
            except Exception:
                self._conn.rollback()

    def commit(self):
        self._conn.commit()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type:
            self._conn.rollback()
        else:
            self._conn.commit()
        self._conn.close()
        return False


def _read_env_file():
    """Legge .env dalla stessa cartella di app.py senza dipendenze esterne."""
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
    if not os.path.exists(env_path):
        return
    with open(env_path, encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                key, val = line.split('=', 1)
                os.environ.setdefault(key.strip(), val.strip())

_read_env_file()


def get_db():
    url = os.environ.get('DATABASE_URL', '')
    if not url:
        raise RuntimeError('DATABASE_URL non impostata. Controlla il file .env')
    conn = psycopg2.connect(url, cursor_factory=psycopg2.extras.RealDictCursor)
    return DbWrapper(conn)


def init_db():
    with get_db() as db:
        db.executescript('''
        CREATE TABLE IF NOT EXISTS utenti (
            id SERIAL PRIMARY KEY,
            email TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            is_admin INTEGER DEFAULT 0,
            created_at TEXT DEFAULT to_char(NOW(), 'YYYY-MM-DD HH24:MI:SS')
        );
        CREATE TABLE IF NOT EXISTS offerte (
            id SERIAL PRIMARY KEY,
            user_id INTEGER NOT NULL,
            nome_offerta TEXT NOT NULL,
            tipo TEXT NOT NULL,
            canale TEXT NOT NULL,
            commodity TEXT NOT NULL,
            tipo_consumo TEXT NOT NULL,
            spread REAL DEFAULT 0,
            quota_fissa REAL DEFAULT 0,
            consumo_medio REAL DEFAULT 0,
            ricorrente_mese REAL DEFAULT 0,
            ricorrente_consumo REAL DEFAULT 0,
            sconto REAL DEFAULT 0,
            stato TEXT DEFAULT 'ATTIVA',
            is_demo INTEGER DEFAULT 0,
            created_at TEXT DEFAULT to_char(NOW(), 'YYYY-MM-DD HH24:MI:SS'),
            FOREIGN KEY(user_id) REFERENCES utenti(id)
        );
        CREATE TABLE IF NOT EXISTS fornitori (
            id SERIAL PRIMARY KEY,
            user_id INTEGER NOT NULL,
            nome TEXT NOT NULL,
            commodity TEXT NOT NULL,
            spread_acquisto REAL DEFAULT 0,
            costo_gestione_pdp REAL DEFAULT 0,
            tipologia_pagamento TEXT DEFAULT 'Open credit',
            is_demo INTEGER DEFAULT 0,
            created_at TEXT DEFAULT to_char(NOW(), 'YYYY-MM-DD HH24:MI:SS'),
            FOREIGN KEY(user_id) REFERENCES utenti(id)
        );
        CREATE TABLE IF NOT EXISTS piani_provvigionali (
            id SERIAL PRIMARY KEY,
            user_id INTEGER NOT NULL,
            nome_piano TEXT NOT NULL,
            gettone_agente REAL DEFAULT 0,
            gettone_sub_agente REAL DEFAULT 0,
            ricorrente_mese_agente REAL DEFAULT 0,
            ricorrente_consumo_agente REAL DEFAULT 0,
            ricorrente_mese_sub_agente REAL DEFAULT 0,
            ricorrente_consumo_sub_agente REAL DEFAULT 0,
            ricorrente_mese_area_manager REAL DEFAULT 0,
            ricorrente_consumo_area_manager REAL DEFAULT 0,
            is_demo INTEGER DEFAULT 0,
            created_at TEXT DEFAULT to_char(NOW(), 'YYYY-MM-DD HH24:MI:SS'),
            FOREIGN KEY(user_id) REFERENCES utenti(id)
        );
        CREATE TABLE IF NOT EXISTS portafogli (
            id SERIAL PRIMARY KEY,
            user_id INTEGER NOT NULL,
            nome TEXT NOT NULL,
            descrizione TEXT,
            is_demo INTEGER DEFAULT 0,
            agente_id INTEGER DEFAULT NULL,
            created_at TEXT DEFAULT to_char(NOW(), 'YYYY-MM-DD HH24:MI:SS'),
            FOREIGN KEY(user_id) REFERENCES utenti(id)
        );
        CREATE TABLE IF NOT EXISTS clienti_portafoglio (
            id SERIAL PRIMARY KEY,
            portafoglio_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            nome_cliente TEXT,
            offerta_id INTEGER,
            piano_id INTEGER,
            consumo_override REAL,
            note TEXT,
            nome_offerta TEXT,
            nome_fornitore TEXT,
            nome_piano TEXT,
            commodity TEXT,
            spread_vendita REAL,
            spread_acquisto REAL,
            quota_fissa REAL,
            costo_gestione_pdp REAL,
            margine_lordo REAL,
            totale_provvigioni REAL,
            margine_netto REAL,
            margine_percentuale REAL,
            created_at TEXT DEFAULT to_char(NOW(), 'YYYY-MM-DD HH24:MI:SS'),
            FOREIGN KEY(portafoglio_id) REFERENCES portafogli(id),
            FOREIGN KEY(user_id) REFERENCES utenti(id)
        );
        CREATE TABLE IF NOT EXISTS agenti (
            id SERIAL PRIMARY KEY,
            user_id INTEGER NOT NULL,
            nome TEXT NOT NULL,
            cognome TEXT NOT NULL DEFAULT '',
            email TEXT DEFAULT '',
            telefono TEXT DEFAULT '',
            ruolo TEXT NOT NULL DEFAULT 'AGENTE',
            zona TEXT DEFAULT '',
            parent_id INTEGER DEFAULT NULL,
            piano_id INTEGER DEFAULT NULL,
            data_attivazione TEXT DEFAULT '',
            stato TEXT DEFAULT 'ATTIVO',
            note TEXT DEFAULT '',
            target_margine_annuo REAL DEFAULT 0,
            target_clienti INTEGER DEFAULT 0,
            created_at TEXT DEFAULT to_char(NOW(), 'YYYY-MM-DD HH24:MI:SS'),
            FOREIGN KEY(user_id) REFERENCES utenti(id)
        );
        CREATE TABLE IF NOT EXISTS simulazioni (
            id SERIAL PRIMARY KEY,
            user_id INTEGER NOT NULL,
            nome_offerta TEXT,
            nome_fornitore TEXT,
            nome_piano TEXT,
            commodity TEXT,
            tipo_consumo TEXT,
            consumo_medio REAL,
            spread_vendita REAL,
            quota_fissa REAL,
            spread_acquisto REAL,
            costo_gestione_pdp REAL,
            margine_spread_annuo REAL,
            margine_qf_annuo REAL,
            margine_lordo REAL,
            provvigione_agente REAL,
            provvigione_sub_agente REAL,
            provvigione_area_manager REAL,
            totale_provvigioni REAL,
            margine_netto REAL,
            margine_percentuale REAL,
            note TEXT,
            is_demo INTEGER DEFAULT 0,
            created_at TEXT DEFAULT to_char(NOW(), 'YYYY-MM-DD HH24:MI:SS'),
            FOREIGN KEY(user_id) REFERENCES utenti(id)
        );
        CREATE TABLE IF NOT EXISTS prezzi_mercato (
            id SERIAL PRIMARY KEY,
            commodity TEXT NOT NULL,
            anno INTEGER NOT NULL,
            mese INTEGER NOT NULL,
            prezzo_mwh REAL NOT NULL,
            fonte TEXT DEFAULT 'GME',
            aggiornato_il TEXT DEFAULT to_char(NOW(), 'YYYY-MM-DD HH24:MI:SS'),
            UNIQUE(commodity, anno, mese)
        )
        ''')
        for email in ADMIN_EMAILS:
            db.execute('UPDATE utenti SET is_admin=1 WHERE lower(email)=?', (email,))
        db.commit()

def seed_demo(admin_user_id):
    """
    Inserisce dati demo realistici per il mercato energia italiano.
    Eseguito una sola volta se il DB è vuoto per quell'utente.
    """
    with get_db() as db:
        if db.execute('SELECT 1 FROM offerte WHERE user_id=? AND is_demo=1', (admin_user_id,)).fetchone():
            return  # demo già presente
        # ── Offerte ────────────────────────────────────────────────────────────
        offerte = [
            # (nome, tipo, canale, commodity, tipo_consumo, spread, qf, consumo_medio, ric_mese, ric_cons)
            ('FLEXI LUCE DOM','Mercato libero','Agente','LUCE','DOMESTICO', 0.0055, 9.50,  2.7,  0, 0),
            ('FLEXI LUCE BIZ','Mercato libero','Agente','LUCE','BUSINESS',  0.0048, 14.00, 18.0, 0, 0),
            ('SMART LUCE IND','Mercato libero','Agente','LUCE','INDUSTRIALE',0.0038, 22.00, 120.0,0, 0),
            ('FLEXI GAS DOM', 'Mercato libero','Agente','GAS', 'DOMESTICO', 0.0280, 7.50,  1400, 0, 0),
            ('SMART GAS BIZ', 'Mercato libero','Agente','GAS', 'BUSINESS',  0.0240, 12.00, 8500, 0, 0),
            ('VERDE LUCE DOM','Green energy',  'Agente','LUCE','DOMESTICO', 0.0062, 10.50, 2.7,  0, 0),
            ('VERDE GAS DOM', 'Green energy',  'Agente','GAS', 'DOMESTICO', 0.0310, 8.00,  1400, 0, 0),
        ]
        offerta_ids = {}
        for o in offerte:
            cur = db.execute('''INSERT INTO offerte
                (user_id,nome_offerta,tipo,canale,commodity,tipo_consumo,spread,quota_fissa,
                 consumo_medio,ricorrente_mese,ricorrente_consumo,stato,is_demo)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,'ATTIVA',1)''', (admin_user_id,)+o)
            offerta_ids[o[0]] = cur.lastrowid
        # ── Fornitori ──────────────────────────────────────────────────────────
        fornitori = [
            # (nome, commodity, spread_acquisto, costo_gestione_pdp, tipologia_pagamento)
            ('Enel Energia',    'LUCE',    0.0018, 3.20, 'Open credit'),
            ('Edison Energia',  'LUCE',    0.0020, 3.50, 'Open credit'),
            ('Eni Plenitude',   'ENTRAMBI',0.0022, 3.80, 'Open credit'),
            ('A2A Energia',     'ENTRAMBI',0.0019, 3.10, 'Anticipo'),
            ('Illumia',         'ENTRAMBI',0.0024, 4.00, 'Anticipo'),
            ('Axpo Italia',     'GAS',     0.0120, 2.90, 'Open credit'),
            ('Sorgenia',        'GAS',     0.0110, 2.70, 'Open credit'),
        ]
        for f in fornitori:
            db.execute('''INSERT INTO fornitori
                (user_id,nome,commodity,spread_acquisto,costo_gestione_pdp,tipologia_pagamento,is_demo)
                VALUES(?,?,?,?,?,?,1)''', (admin_user_id,)+f)
        # ── Piani provvigionali ────────────────────────────────────────────────
        piani = [
            # (nome, gett_ag, gett_sub, ric_m_ag, ric_c_ag, ric_m_sub, ric_c_sub, ric_m_am, ric_c_am)
            ('Base Agente',      80,  0,  0, 0.0010, 0, 0,      0, 0),
            ('Standard Agente', 100, 25,  0, 0.0012, 0, 0.0003, 0, 0),
            ('Premium Agente',  120, 35, 2,  0.0015, 1, 0.0004, 1, 0.0002),
            ('Business Team',   150, 50, 3,  0.0018, 1, 0.0005, 2, 0.0003),
        ]
        piano_ids = {}
        for p in piani:
            cur = db.execute('''INSERT INTO piani_provvigionali
                (user_id,nome_piano,gettone_agente,gettone_sub_agente,
                 ricorrente_mese_agente,ricorrente_consumo_agente,
                 ricorrente_mese_sub_agente,ricorrente_consumo_sub_agente,
                 ricorrente_mese_area_manager,ricorrente_consumo_area_manager,is_demo)
                VALUES(?,?,?,?,?,?,?,?,?,?,1)''', (admin_user_id,)+p)
            piano_ids[p[0]] = cur.lastrowid
        db.commit()
        # ── Simulazioni storiche (ultime 8 settimane) ─────────────────────────
        import random
        from datetime import timedelta
        random.seed(42)
        sim_combos = [
            ('FLEXI LUCE DOM','Enel Energia',  'Standard Agente','LUCE','DOMESTICO',  2.7,  0.0055,9.50, 0.0018,3.20),
            ('FLEXI LUCE BIZ','Edison Energia','Standard Agente','LUCE','BUSINESS',   18.0, 0.0048,14.00,0.0020,3.50),
            ('SMART LUCE IND','A2A Energia',   'Business Team',  'LUCE','INDUSTRIALE',120.0,0.0038,22.00,0.0019,3.10),
            ('FLEXI GAS DOM', 'Eni Plenitude', 'Base Agente',    'GAS', 'DOMESTICO',  1400, 0.0280,7.50, 0.0022,3.80),
            ('SMART GAS BIZ', 'Axpo Italia',   'Premium Agente', 'GAS', 'BUSINESS',   8500, 0.0240,12.00,0.0120,2.90),
            ('VERDE LUCE DOM','A2A Energia',   'Standard Agente','LUCE','DOMESTICO',  2.7,  0.0062,10.50,0.0019,3.10),
            ('VERDE GAS DOM', 'Sorgenia',      'Standard Agente','GAS', 'DOMESTICO',  1400, 0.0310,8.00, 0.0110,2.70),
        ]
        notes_pool = [
            'Cliente acquisito da concorrenza — ex Enel Servizio Elettrico',
            'Rinnovo contrattuale — cliente fidelizzato da 3 anni',
            'Prospect profilato tramite campagna outbound',
            'Azienda manifatturiera — settore metalmeccanico, consumo stabile',
            'Condominio residenziale — accordo con amministratore',
            'PMI nord Italia — referral da agente senior',
            'Cliente referral — segnalazione diretta',
            'Switch da operatore incumbent — offerta competitiva',
            'Accordo quadro multi-sito — primo contratto pilota',
            'Rinegoziazione condizioni — incremento consumo previsto',
            'Nuovo contratto triennale — prezzo fisso',
            'Cliente recuperato dopo scadenza — offerta migliorativa',
        ]
        base_date = datetime.now()
        for week in range(8, 0, -1):
            n_sims = random.randint(3, 6)
            for _ in range(n_sims):
                combo = random.choice(sim_combos)
                (nome_off,nome_for,nome_piano,comm,tipo_cons,
                 cons_base,sv,qf,sa,cpd) = combo
                consumo = round(cons_base * random.uniform(0.7, 1.4), 4)
                pof = offerta_ids.get(nome_off); ppiano = piano_ids.get(nome_piano)
                if not pof or not ppiano: continue
                piano_row = db.execute('SELECT * FROM piani_provvigionali WHERE id=?',(ppiano,)).fetchone()
                if not piano_row: continue
                offerta_v  = {'spread':sv,'quota_fissa':qf,'consumo_medio':cons_base}
                fornitore_v= {'spread_acquisto':sa,'costo_gestione_pdp':cpd,'nome':nome_for}
                r = calcola_simulazione(offerta_v, fornitore_v, dict(piano_row), consumo)
                sim_date = (base_date - timedelta(weeks=week, days=random.randint(0,6))).strftime('%Y-%m-%dT%H:%M')
                db.execute('''INSERT INTO simulazioni
                    (user_id,nome_offerta,nome_fornitore,nome_piano,commodity,tipo_consumo,
                     consumo_medio,spread_vendita,quota_fissa,spread_acquisto,costo_gestione_pdp,
                     margine_spread_annuo,margine_qf_annuo,margine_lordo,provvigione_agente,
                     provvigione_sub_agente,provvigione_area_manager,totale_provvigioni,
                     margine_netto,margine_percentuale,note,created_at,is_demo)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1)''',
                    (admin_user_id,nome_off,nome_for,nome_piano,comm,tipo_cons,
                     consumo,r['spread_vendita'],r['quota_fissa'],r['spread_acquisto'],r['costo_gestione_pdp'],
                     r['margine_spread_annuo'],r['margine_qf_annuo'],r['margine_lordo'],
                     r['provvigione_agente'],r['provvigione_sub_agente'],r['provvigione_area_manager'],
                     r['totale_provvigioni'],r['margine_netto'],r['margine_percentuale'],
                     random.choice(notes_pool), sim_date))
        # ── Portafoglio demo ───────────────────────────────────────────────────
        cur = db.execute("INSERT INTO portafogli(user_id,nome,descrizione,is_demo) VALUES(?,?,?,1)",
            (admin_user_id,'Portafoglio Agente — Q2 2025','Mix residenziale e business: 5 clienti luce, 3 gas. Margine medio positivo, semaforo verde.'))
        pf_id = cur.lastrowid
        clienti_demo = [
            # (nome_cliente, nome_offerta, nome_piano, consumo)
            # Offerta con QF alta abbinata a piani con gettone compatibile col lordo
            ('Rossi Mario',       'VERDE LUCE DOM', 'Base Agente',     3.5),   # dom LUCE, QF alta → netto positivo
            ('Trattoria da Luca', 'FLEXI LUCE BIZ', 'Standard Agente', 22.0),  # BIZ LUCE, consumo realistico
            ('Condominio Via Po', 'FLEXI LUCE BIZ', 'Base Agente',     35.0),  # condo trattato come BIZ, gettone base
            ('Officine Bianchi',  'SMART LUCE IND', 'Standard Agente', 150.0), # ind LUCE, piano standard (no Business Team)
            ('Famiglia Greco',    'FLEXI GAS DOM',  'Base Agente',     2000),  # gas dom, consumo medio-alto
            ('Bar Roma Snc',      'SMART GAS BIZ',  'Standard Agente', 7800),  # gas BIZ
            ('Verdi Costruzioni', 'SMART GAS BIZ',  'Premium Agente',  9200),  # gas BIZ premium
            ('Studio Marini',     'VERDE LUCE DOM', 'Base Agente',     4.0),   # studio, dom LUCE green
        ]
        for (nome_cl, nome_off, nome_piano, consumo) in clienti_demo:
            oid = offerta_ids.get(nome_off); pid = piano_ids.get(nome_piano)
            if not oid or not pid: continue
            offerta_row   = db.execute('SELECT * FROM offerte WHERE id=?',(oid,)).fetchone()
            piano_row     = db.execute('SELECT * FROM piani_provvigionali WHERE id=?',(pid,)).fetchone()
            if not offerta_row or not piano_row: continue
            off_d = dict(offerta_row)
            for_d = get_fornitore_medio(off_d['commodity'])
            if not for_d: continue
            r = calcola_simulazione(off_d, for_d, dict(piano_row), consumo)
            db.execute('''INSERT INTO clienti_portafoglio
                (portafoglio_id,user_id,nome_cliente,offerta_id,piano_id,consumo_override,
                 nome_offerta,nome_fornitore,nome_piano,commodity,
                 spread_vendita,spread_acquisto,quota_fissa,costo_gestione_pdp,
                 margine_lordo,totale_provvigioni,margine_netto,margine_percentuale)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                (pf_id,admin_user_id,nome_cl,oid,pid,consumo,
                 off_d['nome_offerta'],for_d.get('nome','Media'),off_d['nome_offerta'][:10],
                 off_d['commodity'],r['spread_vendita'],r['spread_acquisto'],
                 r['quota_fissa'],r['costo_gestione_pdp'],
                 r['margine_lordo'],r['totale_provvigioni'],r['margine_netto'],r['margine_percentuale']))
        db.commit()


def clear_demo(user_id):
    """Rimuove tutti i dati demo (is_demo=1) per l'utente, senza toccare i dati reali."""
    with get_db() as db:
        # 1. Nomi offerte demo (prima di cancellarle)
        demo_off_names = [r[0] for r in db.execute(
            'SELECT nome_offerta FROM offerte WHERE user_id=? AND is_demo=1', (user_id,)).fetchall()]
        # 2. Id portafogli demo
        demo_pf_ids = [r[0] for r in db.execute(
            'SELECT id FROM portafogli WHERE user_id=? AND is_demo=1', (user_id,)).fetchall()]
        # 3. Cancella clienti portafoglio demo
        if demo_pf_ids:
            ph = ','.join('?' * len(demo_pf_ids))
            db.execute(f'DELETE FROM clienti_portafoglio WHERE portafoglio_id IN ({ph})', demo_pf_ids)
            db.execute(f'DELETE FROM portafogli WHERE id IN ({ph})', demo_pf_ids)
        # 4. Cancella simulazioni demo tramite flag is_demo
        db.execute('DELETE FROM simulazioni WHERE user_id=? AND is_demo=1', (user_id,))
        # 5. Cancella anagrafica demo
        db.execute('DELETE FROM offerte WHERE user_id=? AND is_demo=1', (user_id,))
        db.execute('DELETE FROM fornitori WHERE user_id=? AND is_demo=1', (user_id,))
        db.execute('DELETE FROM piani_provvigionali WHERE user_id=? AND is_demo=1', (user_id,))
        db.commit()


with app.app_context():
    init_db()

# ═══════════════════════════════════════════════════════════════════════════════
# ── PREZZI MERCATO (PUN / PSV) ──────────────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════════

MESI_IT = ['','Gen','Feb','Mar','Apr','Mag','Giu','Lug','Ago','Set','Ott','Nov','Dic']
GAS_MWH_TO_SMC = 0.01

def get_prezzi_ultimi_12_mesi(commodity):
    """Ultimi 12 mesi disponibili, ordinati cronologicamente (più vecchio prima)."""
    with get_db() as db:
        rows = db.execute(
            'SELECT anno,mese,prezzo_mwh,fonte FROM prezzi_mercato '
            'WHERE commodity=? ORDER BY anno DESC,mese DESC LIMIT 12',
            (commodity,)
        ).fetchall()
    return [dict(r) for r in reversed(rows)] if rows else []

def calcola_costo_cliente_12m(sim, prezzi):
    """Stima costo annuo cliente usando prezzi PUN/PSV mensili."""
    if not prezzi:
        return [], 0, 0, 0
    commodity     = sim.get('commodity', 'LUCE')
    spread        = sim.get('spread_vendita', 0) or 0
    qf            = sim.get('quota_fissa', 0) or 0
    consumo_annuo = sim.get('consumo_medio', 0) or 0
    monthly = []
    for p in prezzi:
        pmwh = p.get('prezzo_mwh', 0) or 0
        pu = pmwh * GAS_MWH_TO_SMC if commodity == 'GAS' else pmwh
        cons_mese   = consumo_annuo / 12
        c_energia   = (pu + spread) * cons_mese
        c_mensile   = c_energia + qf
        monthly.append({
            'anno': p['anno'], 'mese': p['mese'],
            'label': f"{MESI_IT[p['mese']]} {p['anno']}",
            'prezzo_mwh': round(pmwh, 2),
            'prezzo_unit': round(pu, 4),
            'costo_energia': round(c_energia, 2),
            'costo_fisso': round(qf, 2),
            'costo_mensile': round(c_mensile, 2),
        })
    n = len(prezzi) or 1
    scale = 12 / n
    tot       = round(sum(m['costo_mensile'] for m in monthly) * scale, 2)
    tot_en    = round(sum(m['costo_energia'] for m in monthly) * scale, 2)
    tot_fisso = round(sum(m['costo_fisso']   for m in monthly) * scale, 2)
    return monthly, tot, tot_en, tot_fisso

# ── Auth ────────────────────────────────────────────────────────────────────────
class User(UserMixin):
    def __init__(self, id, email, is_admin=False):
        self.id       = id
        self.email    = email
        self.is_admin = is_admin

@login_manager.user_loader
def load_user(user_id):
    with get_db() as db:
        u = db.execute('SELECT * FROM utenti WHERE id=?', (user_id,)).fetchone()
        if not u:
            return None
        is_admin = bool(u['is_admin']) or u['email'].lower() in ADMIN_EMAILS
        return User(u['id'], u['email'], is_admin)

# ── Decoratori ──────────────────────────────────────────────────────────────────
def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash('Questa sezione è riservata agli amministratori.', 'warning')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

# ── Helper: filtro utente / admin ───────────────────────────────────────────────
def uid_and():
    if current_user.is_admin:
        return '', []
    return 'AND user_id=?', [current_user.id]

def uid_where():
    if current_user.is_admin:
        return '', []
    return 'WHERE user_id=?', [current_user.id]

# ── Scenario: demo / reale / tutti ──────────────────────────────────────────────
def get_scenario():
    """Restituisce lo scenario attivo: 'all' | 'demo' | 'real'. Default: 'real'."""
    return session.get('scenario', 'real')

def with_scenario(aw, ap):
    """
    Aggiunge il filtro is_demo a una WHERE clause esistente.
    aw: stringa WHERE (es. 'WHERE user_id=?' o '')
    ap: lista di parametri corrispondenti
    Restituisce (nuova_aw, nuovi_ap)
    """
    sc = get_scenario()
    if sc == 'all':
        return aw, list(ap)
    val = 1 if sc == 'demo' else 0
    if aw:
        return aw + ' AND is_demo=?', list(ap) + [val]
    return 'WHERE is_demo=?', [val]

@app.context_processor
def inject_scenario():
    return {'current_scenario': get_scenario()}

@app.route('/toggle-scenario', methods=['POST'])
@login_required
def toggle_scenario():
    sc = request.form.get('scenario', 'all')
    if sc in ('all', 'demo', 'real'):
        session['scenario'] = sc
    return redirect(request.referrer or url_for('dashboard'))

# ── Helper: fornitore medio ─────────────────────────────────────────────────────
def get_fornitore_medio(commodity):
    """
    Calcola il fornitore medio per commodity.
    Admin usa tutti i fornitori di tutti gli utenti.
    Utente normale usa solo i propri — ma siccome non gestisce fornitori,
    l'admin è l'unico a inserirli: tutti vedono sempre la media globale.
    """
    with get_db() as db:
        rows = db.execute(
            'SELECT * FROM fornitori WHERE commodity=? OR commodity="ENTRAMBI"',
            (commodity,)
        ).fetchall()
    if not rows:
        return None
    avg_spread = sum(r['spread_acquisto'] for r in rows) / len(rows)
    avg_costo  = sum(r['costo_gestione_pdp'] for r in rows) / len(rows)
    return {
        'id': 'avg',
        'nome': f'Media {len(rows)} fornitor{"e" if len(rows)==1 else "i"} ({commodity})',
        'commodity': commodity,
        'spread_acquisto': round(avg_spread, 4),
        'costo_gestione_pdp': round(avg_costo, 4),
    }

# ── Calcoli ─────────────────────────────────────────────────────────────────────
def calcola_simulazione(offerta, fornitore, piano, consumo_override=None):
    consumo      = consumo_override if consumo_override is not None else offerta['consumo_medio']
    spread_netto = offerta['spread'] - fornitore['spread_acquisto']
    margine_spread_annuo = spread_netto * consumo
    margine_qf_annuo     = (offerta['quota_fissa'] - fornitore['costo_gestione_pdp']) * 12
    sconto_annuo         = offerta.get('sconto', 0) or 0
    margine_lordo        = margine_spread_annuo + margine_qf_annuo - sconto_annuo

    provv_agente = (piano['gettone_agente']
                    + piano['ricorrente_mese_agente'] * 12
                    + piano['ricorrente_consumo_agente'] * consumo)
    provv_sub    = (piano['gettone_sub_agente']
                    + piano['ricorrente_mese_sub_agente'] * 12
                    + piano['ricorrente_consumo_sub_agente'] * consumo)
    provv_am     = (piano['ricorrente_mese_area_manager'] * 12
                    + piano['ricorrente_consumo_area_manager'] * consumo)
    tot_provv    = provv_agente + provv_sub + provv_am
    margine_netto = margine_lordo - tot_provv
    margine_pct   = (margine_netto / margine_lordo * 100) if margine_lordo != 0 else 0

    ric_cons_tot   = (piano['ricorrente_consumo_agente']
                      + piano['ricorrente_consumo_sub_agente']
                      + piano['ricorrente_consumo_area_manager'])
    gettoni_fissi  = (piano['gettone_agente'] + piano['gettone_sub_agente']
                      + (piano['ricorrente_mese_agente'] + piano['ricorrente_mese_sub_agente']
                         + piano['ricorrente_mese_area_manager']) * 12)
    qf_netta_annua = (offerta['quota_fissa'] - fornitore['costo_gestione_pdp']) * 12
    coeff_consumo  = spread_netto - ric_cons_tot
    if coeff_consumo > 0:
        breakeven = max(0, (gettoni_fissi - qf_netta_annua) / coeff_consumo)
    else:
        breakeven = None

    warnings = []
    if spread_netto < 0:    warnings.append('spread_negativo')
    if margine_lordo < 0:   warnings.append('lordo_negativo')

    return dict(
        consumo_medio            = round(consumo, 4),
        spread_vendita           = offerta['spread'],
        quota_fissa              = offerta['quota_fissa'],
        spread_acquisto          = fornitore['spread_acquisto'],
        costo_gestione_pdp       = fornitore['costo_gestione_pdp'],
        nome_fornitore           = fornitore.get('nome', ''),
        fornitore_medio          = fornitore.get('id') == 'avg',
        margine_spread_annuo     = round(margine_spread_annuo, 2),
        margine_qf_annuo         = round(margine_qf_annuo, 2),
        sconto_annuo             = round(sconto_annuo, 2),
        margine_lordo            = round(margine_lordo, 2),
        provvigione_agente       = round(provv_agente, 2),
        provvigione_sub_agente   = round(provv_sub, 2),
        provvigione_area_manager = round(provv_am, 2),
        totale_provvigioni       = round(tot_provv, 2),
        margine_netto            = round(margine_netto, 2),
        margine_percentuale      = round(margine_pct, 2),
        breakeven                = round(breakeven, 4) if breakeven is not None else None,
        warnings                 = warnings,
    )

# ── Routes ──────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return render_template('landing.html')

@app.route('/landing')
def landing():
    return render_template('landing.html')

def _login_view():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    error = None
    if request.method == 'POST':
        email    = request.form['email'].strip().lower()
        password = request.form['password']
        mode     = request.form.get('mode', 'login')
        with get_db() as db:
            if mode == 'signup':
                if not REGISTRATION_OPEN:
                    error = 'La registrazione pubblica è disabilitata. Contatta un amministratore.'
                elif len(password) < MIN_PASSWORD_LEN:
                    error = f'La password deve essere di almeno {MIN_PASSWORD_LEN} caratteri.'
                elif db.execute('SELECT id FROM utenti WHERE email=?', (email,)).fetchone():
                    error = 'Email già registrata.'
                else:
                    is_admin_new = 1 if email in ADMIN_EMAILS else 0
                    db.execute('INSERT INTO utenti(email,password,is_admin) VALUES(?,?,?)',
                               (email, generate_password_hash(password), is_admin_new))
                    db.commit()
                    u = db.execute('SELECT * FROM utenti WHERE email=?', (email,)).fetchone()
                    login_user(User(u['id'], u['email'], bool(u['is_admin'])), remember=True)
                    seed_demo(u['id'])
                    return redirect(url_for('dashboard'))
            else:
                u = db.execute('SELECT * FROM utenti WHERE email=?', (email,)).fetchone()
                if u and check_password_hash(u['password'], password):
                    is_admin = bool(u['is_admin']) or email in ADMIN_EMAILS
                    login_user(User(u['id'], u['email'], is_admin), remember=True)
                    return redirect(url_for('dashboard'))
                error = 'Email o password non corretti.'
    return render_template('login.html', error=error, registration_open=REGISTRATION_OPEN)

if limiter:
    @app.route('/login', methods=['GET', 'POST'])
    @limiter.limit('10 per minute', methods=['POST'])
    def login():
        return _login_view()
else:
    @app.route('/login', methods=['GET', 'POST'])
    def login():
        return _login_view()

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# ── Dashboard ───────────────────────────────────────────────────────────────────
@app.route('/dashboard')
@login_required
def dashboard():
    aw, ap   = uid_where()
    anda, andp = uid_and()
    sc_aw, sc_ap = with_scenario(aw, ap)
    with get_db() as db:
        sims        = db.execute(f'SELECT * FROM simulazioni {sc_aw} ORDER BY created_at DESC', sc_ap).fetchall()
        n_offerte   = db.execute(f'SELECT COUNT(*) FROM offerte WHERE stato="ATTIVA" {anda}', andp).fetchone()[0]
        n_fornitori = db.execute('SELECT COUNT(*) FROM fornitori').fetchone()[0]  # sempre globale
        n_piani     = db.execute(f'SELECT COUNT(*) FROM piani_provvigionali {aw}', ap).fetchone()[0]

    sims      = [dict(s) for s in sims]
    n_sim     = len(sims)
    avg_netto = round(sum(s['margine_netto'] for s in sims) / n_sim, 2) if n_sim else 0
    avg_lordo = round(sum(s['margine_lordo'] for s in sims) / n_sim, 2) if n_sim else 0
    avg_pct   = round(sum(s['margine_percentuale'] for s in sims if s['margine_percentuale']) / n_sim, 1) if n_sim else 0

    luce_sims = [s for s in sims if s.get('commodity') == 'LUCE']
    gas_sims  = [s for s in sims if s.get('commodity') == 'GAS']
    luce_avg  = round(sum(s['margine_netto'] for s in luce_sims) / len(luce_sims), 2) if luce_sims else 0
    gas_avg   = round(sum(s['margine_netto'] for s in gas_sims)  / len(gas_sims),  2) if gas_sims  else 0

    # ── Ratio netto / provvigioni ─────────────────────────────────────────────
    tot_netto_all = sum(s['margine_netto'] for s in sims)
    tot_provv_all = sum((s['totale_provvigioni'] or 0) for s in sims)
    ratio_netto_provv = round(tot_netto_all / tot_provv_all, 2) if tot_provv_all else None

    top3    = sorted(sims, key=lambda s: s['margine_netto'], reverse=True)[:3]
    recenti = sims[:5]

    # ── Trend settimane (ultimi 8 settimane — solo settimane con dati) ──────────
    from collections import defaultdict
    week_buckets = defaultdict(list)
    now = datetime.now()
    for s in sims:
        try:
            dt = datetime.fromisoformat(s['created_at'])
            delta_days = (now - dt).days
            week_num = min(delta_days // 7, 7)  # 0=questa sett., 7=8 sett. fa
            week_buckets[week_num].append(s['margine_netto'])
        except Exception:
            pass
    # Solo settimane con almeno una simulazione, ordinate dal più vecchio al più recente
    trend_labels  = []
    trend_values  = []
    trend_counts  = []
    for w in range(7, -1, -1):
        vals = week_buckets.get(w, [])
        if not vals:
            continue
        label = 'Questa sett.' if w == 0 else f'-{w}sett.'
        trend_labels.append(label)
        trend_values.append(round(sum(vals) / len(vals), 2))
        trend_counts.append(len(vals))

    # ── Week-over-week delta ──────────────────────────────────────────────────
    this_week_sims = week_buckets.get(0, [])
    last_week_sims = week_buckets.get(1, [])
    delta_n_sim = len(this_week_sims) - len(last_week_sims)
    tw_avg_n = sum(this_week_sims) / len(this_week_sims) if this_week_sims else None
    lw_avg_n = sum(last_week_sims) / len(last_week_sims) if last_week_sims else None
    delta_avg_netto = round(tw_avg_n - lw_avg_n, 2) if (tw_avg_n is not None and lw_avg_n is not None) else None

    # ── Portafogli con semaforo salute ─────────────────────────────────────────
    pf_aw, pf_ap = with_scenario(aw.replace('user_id','p.user_id'), ap)
    with get_db() as db:
        pf_rows = db.execute(
            f'''SELECT p.id, p.nome,
                    COUNT(c.id) AS n_clienti,
                    SUM(c.margine_netto) AS tot_netto,
                    SUM(CASE WHEN c.margine_netto < 0 THEN 1 ELSE 0 END) AS n_negativi
                FROM portafogli p
                LEFT JOIN clienti_portafoglio c ON c.portafoglio_id = p.id
                {pf_aw}
                GROUP BY p.id ORDER BY p.created_at DESC LIMIT 5''', pf_ap
        ).fetchall()
    portafogli_dash = []
    for pf in pf_rows:
        pf_d = dict(pf)
        nc = pf_d['n_clienti'] or 0
        nn = pf_d['n_negativi'] or 0
        pct_neg = round(nn / nc * 100) if nc > 0 else 0
        if pct_neg == 0:         semaforo = 'green'
        elif pct_neg <= 25:      semaforo = 'yellow'
        else:                    semaforo = 'red'
        pf_d['pct_negativi'] = pct_neg
        pf_d['semaforo'] = semaforo
        portafogli_dash.append(pf_d)

    return render_template('dashboard.html',
        n_sim=n_sim, avg_netto=avg_netto, avg_lordo=avg_lordo, avg_pct=avg_pct,
        n_offerte=n_offerte, n_fornitori=n_fornitori, n_piani=n_piani,
        n_luce=len(luce_sims), n_gas=len(gas_sims),
        luce_avg=luce_avg, gas_avg=gas_avg,
        top3=top3, recenti=recenti,
        trend_labels=trend_labels, trend_values=trend_values, trend_counts=trend_counts,
        portafogli_dash=portafogli_dash,
        delta_n_sim=delta_n_sim, delta_avg_netto=delta_avg_netto,
        ratio_netto_provv=ratio_netto_provv,
    )

# ── Admin: gestione utenti ──────────────────────────────────────────────────────
@app.route('/admin/utenti')
@login_required
@admin_required
def admin_utenti():
    with get_db() as db:
        utenti = db.execute('''
            SELECT u.id, u.email, u.is_admin, u.created_at,
                COUNT(DISTINCT o.id)  AS n_offerte,
                COUNT(DISTINCT f.id)  AS n_fornitori,
                COUNT(DISTINCT p.id)  AS n_piani,
                COUNT(DISTINCT s.id)  AS n_simulazioni,
                MAX(s.created_at)     AS ultima_simulazione
            FROM utenti u
            LEFT JOIN offerte o              ON o.user_id = u.id
            LEFT JOIN fornitori f            ON f.user_id = u.id
            LEFT JOIN piani_provvigionali p  ON p.user_id = u.id
            LEFT JOIN simulazioni s          ON s.user_id = u.id
            GROUP BY u.id
            ORDER BY u.created_at DESC
        ''').fetchall()
    return render_template('admin_utenti.html', utenti=[dict(u) for u in utenti])

@app.route('/admin/utenti/toggle-admin/<int:uid>', methods=['POST'])
@login_required
@admin_required
def admin_toggle_admin(uid):
    if uid == current_user.id:
        flash('Non puoi modificare il tuo stesso ruolo.', 'warning')
        return redirect(url_for('admin_utenti'))
    with get_db() as db:
        u = db.execute('SELECT * FROM utenti WHERE id=?', (uid,)).fetchone()
        if u:
            new_val = 0 if u['is_admin'] else 1
            db.execute('UPDATE utenti SET is_admin=? WHERE id=?', (new_val, uid))
            db.commit()
            flash(f'Ruolo aggiornato per {u["email"]}.', 'success')
    return redirect(url_for('admin_utenti'))

@app.route('/admin/utenti/crea', methods=['POST'])
@login_required
@admin_required
def admin_crea_utente():
    email    = request.form.get('email', '').strip().lower()
    password = request.form.get('password', '')
    is_admin = 1 if request.form.get('is_admin') == '1' else 0
    if not email or not password:
        flash('Email e password sono obbligatori.', 'warning')
        return redirect(url_for('admin_utenti'))
    if len(password) < MIN_PASSWORD_LEN:
        flash(f'La password deve essere di almeno {MIN_PASSWORD_LEN} caratteri.', 'warning')
        return redirect(url_for('admin_utenti'))
    with get_db() as db:
        if db.execute('SELECT id FROM utenti WHERE email=?', (email,)).fetchone():
            flash('Email già registrata.', 'warning')
            return redirect(url_for('admin_utenti'))
        cur = db.execute('INSERT INTO utenti(email,password,is_admin) VALUES(?,?,?)',
                         (email, generate_password_hash(password), is_admin))
        db.commit()
        new_id = cur.lastrowid
    seed_demo(new_id)
    flash(f'Account creato per {email}.', 'success')
    return redirect(url_for('admin_utenti'))

@app.route('/admin/utenti/elimina/<int:uid>', methods=['POST'])
@login_required
@admin_required
def admin_elimina_utente(uid):
    if uid == current_user.id:
        flash('Non puoi eliminare il tuo account.', 'warning')
        return redirect(url_for('admin_utenti'))
    with get_db() as db:
        u = db.execute('SELECT email FROM utenti WHERE id=?', (uid,)).fetchone()
        if not u:
            flash('Utente non trovato.', 'warning')
            return redirect(url_for('admin_utenti'))
        email = u['email']
        for tbl in ('simulazioni', 'clienti_portafoglio', 'portafogli',
                    'offerte', 'fornitori', 'piani_provvigionali', 'agenti'):
            db.execute(f'DELETE FROM {tbl} WHERE user_id=?', (uid,))
        db.execute('DELETE FROM utenti WHERE id=?', (uid,))
        db.commit()
    flash(f'Utente {email} eliminato.', 'success')
    return redirect(url_for('admin_utenti'))

@app.route('/admin/reset-demo', methods=['POST'])
@login_required
@admin_required
def admin_reset_demo():
    """Cancella i dati demo e li ricarica freschi, senza toccare i dati reali."""
    clear_demo(current_user.id)
    seed_demo(current_user.id)
    flash('Dati demo ripristinati correttamente.', 'success')
    return redirect(request.referrer or url_for('dashboard'))

# ── Simulatore ──────────────────────────────────────────────────────────────────
@app.route('/simulatore')
@login_required
def simulatore():
    aw, ap = uid_where()
    with get_db() as db:
        offerte = db.execute(
            f'SELECT * FROM offerte {aw} {"AND" if aw else "WHERE"} stato="ATTIVA" ORDER BY nome_offerta', ap
        ).fetchall()
        piani = db.execute(
            f'SELECT * FROM piani_provvigionali {aw} ORDER BY nome_piano', ap
        ).fetchall()
        # Fornitori: solo admin li vede/seleziona
        fornitori = []
        if current_user.is_admin:
            fornitori = db.execute('SELECT * FROM fornitori ORDER BY nome').fetchall()

    return render_template('simulatore.html',
        offerte  =[dict(o) for o in offerte],
        fornitori=[dict(f) for f in fornitori],
        piani    =[dict(p) for p in piani],
        is_admin =current_user.is_admin,
    )

def _get_fornitore_for_api(fornitore_id, offerta):
    """Restituisce il dict fornitore. Non-admin usa sempre la media."""
    # Non-admin: ignora fornitore_id, usa sempre media
    if not current_user.is_admin:
        return get_fornitore_medio(offerta['commodity'])
    # Admin: se non specificato o 'avg', usa media
    if not fornitore_id or str(fornitore_id) == 'avg':
        return get_fornitore_medio(offerta['commodity'])
    with get_db() as db:
        row = db.execute('SELECT * FROM fornitori WHERE id=?', (fornitore_id,)).fetchone()
    return dict(row) if row else None

@app.route('/api/fornitore-medio/<commodity>')
@login_required
def api_fornitore_medio(commodity):
    f = get_fornitore_medio(commodity)
    if not f:
        return jsonify({'error': 'Nessun fornitore disponibile per questa commodity'}), 404
    return jsonify(f)

@app.route('/api/simula', methods=['POST'])
@login_required
def api_simula():
    data   = request.json
    anda, andp = uid_and()
    with get_db() as db:
        offerta = db.execute(f'SELECT * FROM offerte WHERE id=? {anda}', [data['offerta_id']] + andp).fetchone()
        piano   = db.execute(f'SELECT * FROM piani_provvigionali WHERE id=? {anda}', [data['piano_id']] + andp).fetchone()
    if not all([offerta, piano]):
        return jsonify({'error': 'Dati non trovati'}), 404
    offerta_d   = dict(offerta)
    fornitore_d = _get_fornitore_for_api(data.get('fornitore_id'), offerta_d)
    if fornitore_d is None:
        return jsonify({'error': 'Nessun fornitore trovato per questa commodity. Inserire almeno un fornitore.'}), 404
    consumo = float(data['consumo_override']) if data.get('consumo_override') else None
    return jsonify(calcola_simulazione(offerta_d, fornitore_d, dict(piano), consumo))

@app.route('/api/sensitivity', methods=['POST'])
@login_required
def api_sensitivity():
    data   = request.json
    anda, andp = uid_and()
    with get_db() as db:
        offerta = db.execute(f'SELECT * FROM offerte WHERE id=? {anda}', [data['offerta_id']] + andp).fetchone()
        piano   = db.execute(f'SELECT * FROM piani_provvigionali WHERE id=? {anda}', [data['piano_id']] + andp).fetchone()
    if not all([offerta, piano]):
        return jsonify({'error': 'Dati non trovati'}), 404
    offerta_d   = dict(offerta)
    fornitore_d = _get_fornitore_for_api(data.get('fornitore_id'), offerta_d)
    if fornitore_d is None:
        return jsonify({'error': 'Nessun fornitore trovato'}), 404
    piano_d = dict(piano)
    base    = offerta_d['consumo_medio']
    righe   = []
    for s in [0.25, 0.50, 0.75, 1.0, 1.25, 1.50, 1.75]:
        c = round(base * s, 4)
        r = calcola_simulazione(offerta_d, fornitore_d, piano_d, c)
        righe.append({'consumo': c, 'pct_base': int(s * 100),
                      'margine_lordo': r['margine_lordo'],
                      'totale_provvigioni': r['totale_provvigioni'],
                      'margine_netto': r['margine_netto'],
                      'margine_percentuale': r['margine_percentuale']})
    return jsonify(righe)

@app.route('/api/storico')
@login_required
def api_storico():
    commodity = request.args.get('commodity', '')
    limit     = int(request.args.get('limit', 50))
    offset    = int(request.args.get('offset', 0))
    aw, ap    = uid_where()
    sc_aw, sc_ap = with_scenario(aw, ap)
    # Per l'admin, prefissa user_id con s.
    sc_aw_s = sc_aw.replace('user_id', 's.user_id')
    # com_and uses alias 's.' solo per la query admin (che usa FROM simulazioni s)
    com_and_s  = (('AND' if sc_aw_s else 'WHERE') + ' s.commodity=?')  if commodity else ''
    com_and_ns = (('AND' if sc_aw   else 'WHERE') + ' commodity=?')    if commodity else ''
    com_p      = [commodity] if commodity else []

    # Admin: include email utente con JOIN
    if current_user.is_admin:
        select = 'SELECT s.*, u.email AS user_email FROM simulazioni s LEFT JOIN utenti u ON s.user_id=u.id'
        with get_db() as db:
            rows  = db.execute(
                f'{select} {sc_aw_s} {com_and_s} ORDER BY s.created_at DESC LIMIT ? OFFSET ?',
                sc_ap + com_p + [limit, offset]
            ).fetchall()
            total = db.execute(
                f'SELECT COUNT(*) FROM simulazioni s {sc_aw_s} {com_and_s}',
                sc_ap + com_p
            ).fetchone()[0]
    else:
        with get_db() as db:
            rows  = db.execute(
                f'SELECT * FROM simulazioni {sc_aw} {com_and_ns} ORDER BY created_at DESC LIMIT ? OFFSET ?',
                sc_ap + com_p + [limit, offset]
            ).fetchall()
            total = db.execute(
                f'SELECT COUNT(*) FROM simulazioni {sc_aw} {com_and_ns}', sc_ap + com_p
            ).fetchone()[0]

    return jsonify({'rows': [dict(r) for r in rows], 'total': total})

@app.route('/api/storico/elimina/<int:sim_id>', methods=['DELETE'])
@login_required
def api_storico_elimina(sim_id):
    anda, andp = uid_and()
    with get_db() as db:
        db.execute(f'DELETE FROM simulazioni WHERE id=? {anda}', [sim_id] + andp)
        db.commit()
    return jsonify({'ok': True})

@app.route('/api/salva-simulazione', methods=['POST'])
@login_required
def api_salva_simulazione():
    data   = request.json
    anda, andp = uid_and()
    with get_db() as db:
        offerta = db.execute(f'SELECT * FROM offerte WHERE id=? {anda}', [data['offerta_id']] + andp).fetchone()
        piano   = db.execute(f'SELECT * FROM piani_provvigionali WHERE id=? {anda}', [data['piano_id']] + andp).fetchone()
    if not all([offerta, piano]):
        return jsonify({'error': 'Dati non trovati'}), 404
    offerta_d   = dict(offerta)
    fornitore_d = _get_fornitore_for_api(data.get('fornitore_id'), offerta_d)
    if fornitore_d is None:
        return jsonify({'error': 'Nessun fornitore trovato'}), 404
    consumo = float(data['consumo_override']) if data.get('consumo_override') else None
    r = calcola_simulazione(offerta_d, fornitore_d, dict(piano), consumo)
    nome_fornitore = fornitore_d.get('nome', 'Media automatica')
    with get_db() as db:
        db.execute('''INSERT INTO simulazioni
            (user_id,nome_offerta,nome_fornitore,nome_piano,commodity,tipo_consumo,
             consumo_medio,spread_vendita,quota_fissa,spread_acquisto,costo_gestione_pdp,
             margine_spread_annuo,margine_qf_annuo,margine_lordo,provvigione_agente,
             provvigione_sub_agente,provvigione_area_manager,totale_provvigioni,
             margine_netto,margine_percentuale,note)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (current_user.id,
             offerta['nome_offerta'], nome_fornitore, piano['nome_piano'],
             offerta['commodity'], offerta['tipo_consumo'],
             r['consumo_medio'], r['spread_vendita'], r['quota_fissa'],
             r['spread_acquisto'], r['costo_gestione_pdp'],
             r['margine_spread_annuo'], r['margine_qf_annuo'], r['margine_lordo'],
             r['provvigione_agente'], r['provvigione_sub_agente'], r['provvigione_area_manager'],
             r['totale_provvigioni'], r['margine_netto'], r['margine_percentuale'],
             data.get('note', '')))
        db.commit()
    return jsonify({'ok': True})

# ── Offerte CRUD ────────────────────────────────────────────────────────────────
@app.route('/offerte')
@login_required
def offerte():
    aw, ap = uid_where()
    with get_db() as db:
        rows = db.execute(f'SELECT * FROM offerte {aw} ORDER BY created_at DESC', ap).fetchall()
    return render_template('offerte.html', offerte=[dict(r) for r in rows])

@app.route('/offerte/salva', methods=['POST'])
@login_required
def offerta_salva():
    f      = request.form
    consumo = float(f.get('consumo_medio') or 0)
    if consumo == 0:
        consumo = CONSUMI_DEFAULT.get((f['tipo_consumo'], f['commodity']), 0)
    row_id = f.get('id')
    with get_db() as db:
        sconto_val = float(f.get('sconto', 0) or 0)
        if row_id:
            db.execute('''UPDATE offerte SET nome_offerta=?,tipo=?,canale=?,commodity=?,tipo_consumo=?,
                spread=?,quota_fissa=?,consumo_medio=?,ricorrente_mese=?,ricorrente_consumo=?,sconto=?,stato=?
                WHERE id=? AND user_id=?''',
                (f['nome_offerta'], f['tipo'], f['canale'], f['commodity'], f['tipo_consumo'],
                 float(f.get('spread',0)), float(f.get('quota_fissa',0)), consumo,
                 float(f.get('ricorrente_mese',0)), float(f.get('ricorrente_consumo',0)),
                 sconto_val, f['stato'], row_id, current_user.id))
            flash('Offerta aggiornata con successo.', 'success')
        else:
            db.execute('''INSERT INTO offerte(user_id,nome_offerta,tipo,canale,commodity,tipo_consumo,
                spread,quota_fissa,consumo_medio,ricorrente_mese,ricorrente_consumo,sconto,stato)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                (current_user.id, f['nome_offerta'], f['tipo'], f['canale'], f['commodity'],
                 f['tipo_consumo'], float(f.get('spread',0)), float(f.get('quota_fissa',0)),
                 consumo, float(f.get('ricorrente_mese',0)), float(f.get('ricorrente_consumo',0)),
                 sconto_val, f['stato']))
            flash('Offerta creata con successo.', 'success')
        db.commit()
    return redirect(url_for('offerte'))

@app.route('/offerte/elimina/<int:row_id>', methods=['POST'])
@login_required
def offerta_elimina(row_id):
    anda, andp = uid_and()
    with get_db() as db:
        db.execute(f'DELETE FROM offerte WHERE id=? {anda}', [row_id] + andp)
        db.commit()
    flash('Offerta eliminata.', 'info')
    return redirect(url_for('offerte'))

# ── Fornitori CRUD — solo admin ─────────────────────────────────────────────────
@app.route('/fornitori')
@login_required
@admin_required
def fornitori():
    with get_db() as db:
        rows = db.execute('SELECT * FROM fornitori ORDER BY created_at DESC').fetchall()
    return render_template('fornitori.html', fornitori=[dict(r) for r in rows])

@app.route('/fornitori/salva', methods=['POST'])
@login_required
@admin_required
def fornitore_salva():
    f      = request.form
    row_id = f.get('id')
    with get_db() as db:
        if row_id:
            db.execute('''UPDATE fornitori SET nome=?,commodity=?,spread_acquisto=?,
                costo_gestione_pdp=?,tipologia_pagamento=? WHERE id=?''',
                (f['nome'], f['commodity'], float(f.get('spread_acquisto',0)),
                 float(f.get('costo_gestione_pdp',0)), f['tipologia_pagamento'], row_id))
            flash('Fornitore aggiornato con successo.', 'success')
        else:
            db.execute('''INSERT INTO fornitori(user_id,nome,commodity,spread_acquisto,costo_gestione_pdp,tipologia_pagamento)
                VALUES(?,?,?,?,?,?)''',
                (current_user.id, f['nome'], f['commodity'],
                 float(f.get('spread_acquisto',0)), float(f.get('costo_gestione_pdp',0)),
                 f['tipologia_pagamento']))
            flash('Fornitore creato con successo.', 'success')
        db.commit()
    return redirect(url_for('fornitori'))

@app.route('/fornitori/elimina/<int:row_id>', methods=['POST'])
@login_required
@admin_required
def fornitore_elimina(row_id):
    with get_db() as db:
        db.execute('DELETE FROM fornitori WHERE id=?', (row_id,))
        db.commit()
    flash('Fornitore eliminato.', 'info')
    return redirect(url_for('fornitori'))

# ── Piani Provvigionali CRUD ────────────────────────────────────────────────────
@app.route('/provvigioni')
@login_required
def provvigioni():
    aw, ap = uid_where()
    with get_db() as db:
        rows = db.execute(f'SELECT * FROM piani_provvigionali {aw} ORDER BY created_at DESC', ap).fetchall()
    return render_template('provvigioni.html', piani=[dict(r) for r in rows])

@app.route('/provvigioni/salva', methods=['POST'])
@login_required
def provvigione_salva():
    f      = request.form
    row_id = f.get('id')
    vals = [f['nome_piano'],
            float(f.get('gettone_agente',0)), float(f.get('gettone_sub_agente',0)),
            float(f.get('ricorrente_mese_agente',0)), float(f.get('ricorrente_consumo_agente',0)),
            float(f.get('ricorrente_mese_sub_agente',0)), float(f.get('ricorrente_consumo_sub_agente',0)),
            float(f.get('ricorrente_mese_area_manager',0)), float(f.get('ricorrente_consumo_area_manager',0))]
    with get_db() as db:
        if row_id:
            db.execute('''UPDATE piani_provvigionali SET nome_piano=?,gettone_agente=?,gettone_sub_agente=?,
                ricorrente_mese_agente=?,ricorrente_consumo_agente=?,ricorrente_mese_sub_agente=?,
                ricorrente_consumo_sub_agente=?,ricorrente_mese_area_manager=?,ricorrente_consumo_area_manager=?
                WHERE id=? AND user_id=?''', vals + [row_id, current_user.id])
            flash('Piano aggiornato con successo.', 'success')
        else:
            db.execute('''INSERT INTO piani_provvigionali(user_id,nome_piano,gettone_agente,gettone_sub_agente,
                ricorrente_mese_agente,ricorrente_consumo_agente,ricorrente_mese_sub_agente,
                ricorrente_consumo_sub_agente,ricorrente_mese_area_manager,ricorrente_consumo_area_manager)
                VALUES(?,?,?,?,?,?,?,?,?,?)''', [current_user.id] + vals)
            flash('Piano creato con successo.', 'success')
        db.commit()
    return redirect(url_for('provvigioni'))

@app.route('/provvigioni/elimina/<int:row_id>', methods=['POST'])
@login_required
def provvigione_elimina(row_id):
    anda, andp = uid_and()
    with get_db() as db:
        db.execute(f'DELETE FROM piani_provvigionali WHERE id=? {anda}', [row_id] + andp)
        db.commit()
    flash('Piano eliminato.', 'info')
    return redirect(url_for('provvigioni'))

# ── Export Excel ────────────────────────────────────────────────────────────────
def _stile_header(ws, row, col_defs):
    hf = PatternFill('solid', fgColor='1a1f2e')
    hfont = Font(bold=True, color='7c9dff', size=10)
    ha = Alignment(horizontal='center', vertical='center')
    hb = Border(bottom=Side(style='thin', color='30363d'))
    for col, (testo, width) in enumerate(col_defs, 1):
        c = ws.cell(row=row, column=col, value=testo)
        c.font=hfont; c.fill=hf; c.alignment=ha; c.border=hb
        ws.column_dimensions[c.column_letter].width = width
    ws.row_dimensions[row].height = 20

def _stile_riga(ws, row_i, num_cols, alternate=False):
    fill = PatternFill('solid', fgColor='1a1f2e' if alternate else '161b22')
    for col in range(1, num_cols+1):
        c = ws.cell(row=row_i, column=col)
        c.fill=fill; c.font=Font(color='e6edf3',size=10)
        c.alignment=Alignment(vertical='center')
        c.border=Border(bottom=Side(style='thin',color='21262d'))

@app.route('/export/simulazioni')
@login_required
def export_simulazioni():
    aw, ap = uid_where()
    with get_db() as db:
        rows = db.execute(f'SELECT * FROM simulazioni {aw} ORDER BY created_at DESC', ap).fetchall()
    wb = Workbook(); ws = wb.active; ws.title = 'Simulazioni'
    col_defs = [('Data',18),('Offerta',22),('Fornitore',18),('Piano',18),
                ('Commodity',12),('Tipo Consumo',14),('Consumo Medio',16),
                ('Spread Vendita',16),('Spread Acquisto',16),('Quota Fissa',14),
                ('M. Spread Annuo',18),('M. QF Annuo',14),('M. Lordo',14),
                ('Provv. Agente',14),('Provv. Sub-Ag.',14),('Provv. AM',12),
                ('Tot. Provvigioni',16),('M. Netto',14),('M. %',10),('Note',24)]
    _stile_header(ws, 1, col_defs)
    keys = ['created_at','nome_offerta','nome_fornitore','nome_piano','commodity','tipo_consumo',
            'consumo_medio','spread_vendita','spread_acquisto','quota_fissa','margine_spread_annuo',
            'margine_qf_annuo','margine_lordo','provvigione_agente','provvigione_sub_agente',
            'provvigione_area_manager','totale_provvigioni','margine_netto','margine_percentuale','note']
    for row_i, s in enumerate(rows, 2):
        _stile_riga(ws, row_i, len(col_defs), alternate=(row_i%2==0))
        for col, k in enumerate(keys, 1):
            v = s[k]
            if k=='margine_percentuale' and v is not None: v=f'{v:.1f}%'
            c = ws.cell(row=row_i, column=col, value=v)
            if k=='margine_netto':
                c.font=Font(color='3fb950' if (v or 0)>=0 else 'f85149',size=10,bold=True)
    ws.freeze_panes='A2'
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    fname=f"simulazioni_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/export/offerte')
@login_required
def export_offerte():
    aw, ap = uid_where()
    with get_db() as db:
        rows = db.execute(f'SELECT * FROM offerte {aw} ORDER BY nome_offerta', ap).fetchall()
    wb=Workbook(); ws=wb.active; ws.title='Offerte'
    col_defs=[('ID',8),('Nome Offerta',22),('Tipo',14),('Canale',14),
              ('Commodity',12),('Tipo Consumo',14),('Spread',14),
              ('Quota Fissa',14),('Consumo Medio',16),
              ('Ric. €/mese',14),('Ric. €/consumo',14),('Sconto €/anno',14),('Stato',12),('Creato il',18)]
    _stile_header(ws, 1, col_defs)
    keys=['id','nome_offerta','tipo','canale','commodity','tipo_consumo',
          'spread','quota_fissa','consumo_medio','ricorrente_mese','ricorrente_consumo','sconto','stato','created_at']
    for row_i, r in enumerate(rows, 2):
        _stile_riga(ws, row_i, len(col_defs), alternate=(row_i%2==0))
        for col, k in enumerate(keys, 1):
            ws.cell(row=row_i, column=col, value=r[k])
    ws.freeze_panes='A2'
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='offerte.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/export/simulazione/<int:sim_id>')
@login_required
def export_simulazione_singola(sim_id):
    # ── Carica simulazione (con controllo ownership) ────────────────────────────
    anda, andp = uid_and()
    with get_db() as db:
        sim = db.execute(
            f'SELECT * FROM simulazioni WHERE id=? {anda}', [sim_id] + andp
        ).fetchone()
    if not sim:
        abort(404)
    sim = dict(sim)

    # ── Ricostruisci offerta/fornitore virtuali dai valori salvati ──────────────
    offerta_v   = {'spread': sim['spread_vendita'], 'quota_fissa': sim['quota_fissa'],
                   'consumo_medio': sim['consumo_medio']}
    fornitore_v = {'spread_acquisto': sim['spread_acquisto'],
                   'costo_gestione_pdp': sim['costo_gestione_pdp'],
                   'nome': sim['nome_fornitore']}

    # ── Cerca piano per sensitivity (best effort: lookup per nome) ──────────────
    piano_d = None
    if sim.get('nome_piano'):
        aw2, ap2 = uid_where()
        sep = 'AND' if aw2 else 'WHERE'
        with get_db() as db:
            pr = db.execute(
                f'SELECT * FROM piani_provvigionali {aw2} {sep} nome_piano=?',
                ap2 + [sim['nome_piano']]
            ).fetchone()
        if pr:
            piano_d = dict(pr)

    # ── Calcola sensitivity 7 step ──────────────────────────────────────────────
    base = sim['consumo_medio']
    sensitivity = []
    for pct in [25, 50, 75, 100, 125, 150, 175]:
        c = round(base * pct / 100, 4)
        if piano_d:
            r = calcola_simulazione(offerta_v, fornitore_v, piano_d, c)
        else:
            sn = sim['spread_vendita'] - sim['spread_acquisto']
            msa = round(sn * c, 2)
            qfa = round((sim['quota_fissa'] - sim['costo_gestione_pdp']) * 12, 2)
            ml  = round(msa + qfa, 2)
            pp  = sim['totale_provvigioni']   # stima fissa (piano non trovato)
            mn  = round(ml - pp, 2)
            r   = dict(consumo_medio=c, margine_spread_annuo=msa, margine_qf_annuo=qfa,
                       margine_lordo=ml, totale_provvigioni=pp,
                       provvigione_agente=sim['provvigione_agente'],
                       provvigione_sub_agente=sim['provvigione_sub_agente'],
                       provvigione_area_manager=sim['provvigione_area_manager'],
                       margine_netto=mn,
                       margine_percentuale=round(mn/ml*100,2) if ml else 0)
        sensitivity.append({'pct': pct, 'consumo': c, **r})

    # ── Breakeven ───────────────────────────────────────────────────────────────
    breakeven_val = None
    if piano_d:
        sn  = sim['spread_vendita'] - sim['spread_acquisto']
        rct = (piano_d['ricorrente_consumo_agente'] + piano_d['ricorrente_consumo_sub_agente']
               + piano_d['ricorrente_consumo_area_manager'])
        gf  = (piano_d['gettone_agente'] + piano_d['gettone_sub_agente']
               + (piano_d['ricorrente_mese_agente'] + piano_d['ricorrente_mese_sub_agente']
                  + piano_d['ricorrente_mese_area_manager']) * 12)
        qfn = (sim['quota_fissa'] - sim['costo_gestione_pdp']) * 12
        coef = sn - rct
        if coef > 0:
            breakeven_val = round(max(0, (gf - qfn) / coef), 4)

    # ── Palette colori ──────────────────────────────────────────────────────────
    BG0 = '0d1117'; BG1 = '161b22'; BG2 = '1a1f2e'
    FG  = 'e6edf3'; FGA = '7c9dff'; FGM = '8b949e'
    GRN = '3fb950'; RED = 'f85149'; ORG = 'f0883e'

    def _c(ws, row, col, val, bold=False, color=FG, bg=None, align='left',
           size=10, fmt=None, italic=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(color=color, bold=bold, size=size, italic=italic)
        c.alignment = Alignment(horizontal=align, vertical='center')
        if bg:
            c.fill = PatternFill('solid', fgColor=bg)
        if fmt:
            c.number_format = fmt
        return c

    def _sec(ws, row, label, ncol=5):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
        c = ws.cell(row=row, column=1, value=f'  {label}')
        c.font      = Font(color=FGA, bold=True, size=9)
        c.fill      = PatternFill('solid', fgColor=BG0)
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 15

    def _kv(ws, row, label, val, vc=None, fmt=None, bold=False):
        _c(ws, row, 1, label, color=FGM, size=9)
        _c(ws, row, 2, val, bold=bold, color=vc or FG, size=10, fmt=fmt,
           bg=BG1 if row % 2 == 0 else None)
        ws.row_dimensions[row].height = 16

    wb = Workbook()

    # ════════════════════════════════════════════════════════════════════════════
    # SHEET 1 — Riepilogo
    # ════════════════════════════════════════════════════════════════════════════
    ws1 = wb.active; ws1.title = 'Riepilogo'
    ws1.sheet_view.showGridLines = False
    unit = 'MWh' if sim.get('commodity') == 'LUCE' else 'smc'

    for col, w in [('A',30),('B',18),('C',3),('D',28),('E',18)]:
        ws1.column_dimensions[col].width = w

    # Title
    ws1.merge_cells('A1:E1')
    t = ws1.cell(row=1, column=1,
                 value=f'ANALISI SIMULAZIONE  —  {sim.get("nome_offerta") or ""}')
    t.font      = Font(color=FGA, bold=True, size=14)
    t.fill      = PatternFill('solid', fgColor=BG0)
    t.alignment = Alignment(horizontal='left', vertical='center')
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells('A2:E2')
    data_str = (sim.get('created_at') or '')[:16].replace('T', ' ')
    sub = ws1.cell(row=2, column=1,
                   value=f'Generato il {datetime.now().strftime("%d/%m/%Y %H:%M")}   •   Simulazione del {data_str}')
    sub.font      = Font(color=FGM, size=9, italic=True)
    sub.fill      = PatternFill('solid', fgColor=BG0)
    sub.alignment = Alignment(horizontal='left', vertical='center')
    ws1.row_dimensions[2].height = 14
    ws1.row_dimensions[3].height = 6

    # ── Colonna sinistra: dati + parametri ──────────────────────────────────────
    r = 4
    _sec(ws1, r, 'DATI SIMULAZIONE', ncol=2); r += 1
    _kv(ws1, r, 'Offerta (CTE)',        sim.get('nome_offerta') or '—', bold=True); r+=1
    _kv(ws1, r, 'Fornitore',            sim.get('nome_fornitore') or '—'); r+=1
    _kv(ws1, r, 'Piano provvigionale',  sim.get('nome_piano') or '—'); r+=1
    _kv(ws1, r, 'Commodity',            sim.get('commodity') or '—'); r+=1
    _kv(ws1, r, 'Tipo consumo',         sim.get('tipo_consumo') or '—'); r+=1
    _kv(ws1, r, f'Consumo ({unit})',    sim.get('consumo_medio'), fmt='#,##0.0000'); r+=1
    _kv(ws1, r, 'Data simulazione',     data_str); r+=1
    _kv(ws1, r, 'Note',                 sim.get('note') or '—'); r+=1

    ws1.row_dimensions[r].height = 8; r+=1

    _sec(ws1, r, 'PARAMETRI TECNICI', ncol=2); r += 1
    sn_val = round(sim['spread_vendita'] - sim['spread_acquisto'], 4)
    _kv(ws1, r, 'Spread vendita (€/unità)',    sim['spread_vendita'],    fmt='€#,##0.0000'); r+=1
    _kv(ws1, r, 'Spread acquisto (€/unità)',   sim['spread_acquisto'],   fmt='€#,##0.0000'); r+=1
    _kv(ws1, r, 'Spread netto (€/unità)',      sn_val, fmt='€#,##0.0000',
        vc=GRN if sn_val >= 0 else RED, bold=True); r+=1
    _kv(ws1, r, 'Quota fissa vendita (€/mese)',  sim['quota_fissa'],       fmt='€#,##0.00'); r+=1
    _kv(ws1, r, 'Costo gestione PDP (€/mese)',   sim['costo_gestione_pdp'],fmt='€#,##0.00'); r+=1

    ws1.row_dimensions[r].height = 8; r+=1
    if breakeven_val is not None:
        _sec(ws1, r, 'BREAKEVEN', ncol=2); r+=1
        bv_color = GRN if (sim['consumo_medio'] or 0) >= breakeven_val else RED
        _kv(ws1, r, f'Consumo di breakeven ({unit})', breakeven_val,
            fmt='#,##0.00', vc=bv_color, bold=True); r+=1
        status = '✅ Sopra il breakeven' if (sim['consumo_medio'] or 0) >= breakeven_val else '⚠️ Sotto il breakeven'
        _kv(ws1, r, 'Stato',  status, vc=bv_color); r+=1

    # ── Colonna destra: composizione margini (dati + grafico) ───────────────────
    rr = 4
    _sec(ws1, rr, 'COMPOSIZIONE MARGINI', ncol=2)

    # Intestazioni dati grafico (col D=4, E=5)
    ws1.row_dimensions[rr].height = 15; rr+=1
    chart_labels = ['M. Spread Annuo','M. QF Annuo','Provv. Agente',
                    'Provv. Sub-Agente','Provv. Area Mgr','Margine Netto']
    chart_values = [sim['margine_spread_annuo'], sim['margine_qf_annuo'],
                    -sim['provvigione_agente'], -sim['provvigione_sub_agente'],
                    -sim['provvigione_area_manager'], sim['margine_netto']]
    v_colors = [FGA, FGA, RED, RED, ORG, GRN if sim['margine_netto'] >= 0 else RED]

    for i, (lbl, val) in enumerate(zip(chart_labels, chart_values)):
        row_i = rr + i
        _c(ws1, row_i, 4, lbl, color=FGM, size=9)
        _c(ws1, row_i, 5, val, bold=(i==5), color=v_colors[i], fmt='€#,##0.00',
           bg=BG1 if row_i % 2 == 0 else None, align='right')
        ws1.row_dimensions[row_i].height = 16

    # Margine Lordo (sotto l'elenco, come totale separato)
    row_lordo = rr + 6
    ws1.row_dimensions[row_lordo].height = 8
    row_lordo += 1
    _c(ws1, row_lordo, 4, 'Margine Lordo',  bold=True, color=FG, size=10)
    _c(ws1, row_lordo, 5, sim['margine_lordo'], bold=True, color=FGA, fmt='€#,##0.00', align='right')
    ws1.row_dimensions[row_lordo].height = 18
    row_lordo += 1
    _c(ws1, row_lordo, 4, 'Tot. Provvigioni', color=FGM, size=9)
    _c(ws1, row_lordo, 5, sim['totale_provvigioni'], color=RED, fmt='€#,##0.00', align='right')
    ws1.row_dimensions[row_lordo].height = 16
    row_lordo += 1
    _c(ws1, row_lordo, 4, 'Margine Netto', bold=True, color=FG, size=12)
    mn_col = GRN if sim['margine_netto'] >= 0 else RED
    _c(ws1, row_lordo, 5, sim['margine_netto'], bold=True, color=mn_col,
       fmt='€#,##0.00', align='right', size=12)
    ws1.row_dimensions[row_lordo].height = 20
    row_lordo += 1
    _c(ws1, row_lordo, 4, 'Margine %', color=FGM, size=9)
    _c(ws1, row_lordo, 5, f'{sim["margine_percentuale"]:.1f}%' if sim["margine_percentuale"] else '—',
       color=mn_col, align='right')
    ws1.row_dimensions[row_lordo].height = 16

    # BarChart composizione
    bc = BarChart()
    bc.type       = 'col'
    bc.grouping   = 'clustered'
    bc.title      = None
    bc.y_axis.title = '€'
    bc.x_axis.title = None
    bc.legend     = None
    bc.style      = 10
    bc.width      = 16; bc.height = 11
    data_ref  = Reference(ws1, min_col=5, max_col=5, min_row=rr, max_row=rr+5)
    cats_ref  = Reference(ws1, min_col=4, max_col=4, min_row=rr, max_row=rr+5)
    bc.add_data(data_ref, titles_from_data=False)
    bc.set_categories(cats_ref)
    bc.series[0].graphicalProperties.solidFill = '4f6ef7'
    bc.anchor = 'D18'
    ws1.add_chart(bc)

    # ════════════════════════════════════════════════════════════════════════════
    # SHEET 2 — Analisi Sensitività
    # ════════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Analisi Sensitività')
    ws2.sheet_view.showGridLines = False

    for col, w in [('A',14),('B',16),('C',18),('D',18),('E',18),('F',16),('G',12)]:
        ws2.column_dimensions[col].width = w

    ws2.merge_cells('A1:G1')
    t2 = ws2.cell(row=1, column=1,
                  value=f'ANALISI DI SENSITIVITÀ — {sim.get("nome_offerta") or ""}  —  Piano: {sim.get("nome_piano") or "—"}')
    t2.font = Font(color=FGA, bold=True, size=12)
    t2.fill = PatternFill('solid', fgColor=BG0)
    t2.alignment = Alignment(horizontal='left', vertical='center')
    ws2.row_dimensions[1].height = 26

    ws2.row_dimensions[2].height = 6
    if not piano_d:
        ws2.merge_cells('A2:G2')
        note_c = ws2.cell(row=2, column=1,
                          value='⚠️  Piano provvigionale non trovato nel DB: provvigioni stimate al valore base.')
        note_c.font = Font(color=ORG, size=9, italic=True)
        note_c.fill = PatternFill('solid', fgColor=BG0)
        ws2.row_dimensions[2].height = 14

    # Header tabella
    cols2 = [f'% Base', f'Consumo ({unit})',
             'M. Spread (€)', 'M. QF (€)', 'M. Lordo (€)',
             'Provv. Tot. (€)', 'M. Netto (€)', 'M. %']
    widths2 = [10, 16, 18, 16, 16, 16, 16, 10]
    for i, (lbl, w) in enumerate(zip(cols2, widths2), 1):
        ws2.column_dimensions[chr(64+i)].width = w
        c2 = ws2.cell(row=3, column=i, value=lbl)
        c2.font      = Font(color=FGA, bold=True, size=9)
        c2.fill      = PatternFill('solid', fgColor=BG2)
        c2.alignment = Alignment(horizontal='center', vertical='center')
        c2.border    = Border(bottom=Side(style='thin', color='30363d'))
    ws2.row_dimensions[3].height = 18

    for ri, s in enumerate(sensitivity, 4):
        alt  = ri % 2 == 0
        bg_r = BG1 if alt else None
        is_base = s['pct'] == 100
        bg_b = '1a2f3a' if is_base else bg_r
        mn_c = GRN if s['margine_netto'] >= 0 else RED

        def sc(col, val, fmt=None, color=FG, bold=False):
            c2 = ws2.cell(row=ri, column=col, value=val)
            c2.font      = Font(color=color, bold=bold or is_base, size=10)
            c2.alignment = Alignment(horizontal='center' if col==1 else 'right', vertical='center')
            c2.fill      = PatternFill('solid', fgColor=bg_b) if bg_b else PatternFill()
            if fmt: c2.number_format = fmt
            return c2

        sc(1, f'{s["pct"]}%',                   color=FGA if is_base else FGM)
        sc(2, s['consumo'],                      fmt='#,##0.0000')
        sc(3, s['margine_spread_annuo'],          fmt='€#,##0.00')
        sc(4, s['margine_qf_annuo'],              fmt='€#,##0.00')
        sc(5, s['margine_lordo'],                 fmt='€#,##0.00', bold=True)
        sc(6, s['totale_provvigioni'],            fmt='€#,##0.00', color=RED)
        sc(7, s['margine_netto'],                 fmt='€#,##0.00', color=mn_c, bold=True)
        pct_str = f'{s["margine_percentuale"]:.1f}%' if s.get("margine_percentuale") is not None else '—'
        sc(8, pct_str, color=mn_c)
        ws2.row_dimensions[ri].height = 17

    # LineChart sensitività
    lc = LineChart()
    lc.title   = None
    lc.style   = 10
    lc.y_axis.title = '€'
    lc.x_axis.title = f'Consumo ({unit})'
    lc.width   = 20; lc.height = 12

    lordo_ref = Reference(ws2, min_col=5, max_col=5, min_row=3, max_row=10)
    prov_ref  = Reference(ws2, min_col=6, max_col=6, min_row=3, max_row=10)
    netto_ref = Reference(ws2, min_col=7, max_col=7, min_row=3, max_row=10)
    cats2     = Reference(ws2, min_col=2, max_col=2, min_row=4, max_row=10)

    lc.add_data(lordo_ref, titles_from_data=True)
    lc.add_data(prov_ref,  titles_from_data=True)
    lc.add_data(netto_ref, titles_from_data=True)
    lc.set_categories(cats2)

    lc.series[0].graphicalProperties.line.solidFill = '4f6ef7'
    lc.series[0].graphicalProperties.line.width = 20000
    lc.series[1].graphicalProperties.line.solidFill = 'f85149'
    lc.series[1].graphicalProperties.line.width = 15000
    lc.series[1].graphicalProperties.line.dashDot = 'dash'
    lc.series[2].graphicalProperties.line.solidFill = '3fb950'
    lc.series[2].graphicalProperties.line.width = 25000

    lc.anchor = 'A12'
    ws2.add_chart(lc)

    # ── Invia file ───────────────────────────────────────────────────────────────
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    offerta_safe = (sim.get('nome_offerta') or 'sim').replace(' ', '_')[:20]
    fname = f'sim_{sim_id}_{offerta_safe}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ══════════════════════════════════════════════════════════════════════════════
# PORTAFOGLI
# ══════════════════════════════════════════════════════════════════════════════

def _calcola_cliente(offerta_id, piano_id, consumo_override):
    """Calcola i margini per un singolo cliente. Ritorna dict o None."""
    aw, ap = uid_and()
    sep = 'AND' if aw else 'WHERE'
    with get_db() as db:
        offerta = db.execute(
            f'SELECT * FROM offerte WHERE id=? {aw}',
            [offerta_id] + ap
        ).fetchone()
        if not offerta:
            offerta = db.execute('SELECT * FROM offerte WHERE id=?', (offerta_id,)).fetchone()
        piano = db.execute(
            f'SELECT * FROM piani_provvigionali WHERE id=? {aw}', [piano_id] + ap
        ).fetchone()
        if not piano:
            piano = db.execute('SELECT * FROM piani_provvigionali WHERE id=?', (piano_id,)).fetchone()
    if not offerta or not piano:
        return None
    offerta_d = dict(offerta)
    piano_d   = dict(piano)
    fornitore_d = _get_fornitore_for_api(None, offerta_d)
    if not fornitore_d:
        return None
    r = calcola_simulazione(offerta_d, fornitore_d, piano_d, consumo_override)
    r['nome_offerta']   = offerta_d['nome_offerta']
    r['nome_fornitore'] = fornitore_d.get('nome', 'Media automatica')
    r['nome_piano']     = piano_d['nome_piano']
    r['commodity']      = offerta_d['commodity']
    return r


@app.route('/portafogli')
@login_required
def portafogli():
    aw, ap = uid_where()
    pf_aw, pf_ap = with_scenario(aw.replace('user_id','p.user_id'), ap)
    with get_db() as db:
        pflist = db.execute(
            f'''SELECT p.*,
                    COUNT(c.id) AS n_clienti,
                    SUM(c.margine_netto) AS tot_netto,
                    SUM(c.margine_lordo) AS tot_lordo
                FROM portafogli p
                LEFT JOIN clienti_portafoglio c ON c.portafoglio_id = p.id
                {pf_aw}
                GROUP BY p.id ORDER BY p.created_at DESC''', pf_ap
        ).fetchall()
    return render_template('portafogli.html', portafogli=[dict(p) for p in pflist])


@app.route('/portafogli/crea', methods=['POST'])
@login_required
def portafoglio_crea():
    nome = request.form.get('nome', '').strip()
    desc = request.form.get('descrizione', '').strip()
    if not nome:
        flash('Inserisci un nome per il portafoglio.', 'warning')
        return redirect(url_for('portafogli'))
    with get_db() as db:
        cur = db.execute(
            'INSERT INTO portafogli(user_id,nome,descrizione) VALUES(?,?,?)',
            (current_user.id, nome, desc)
        )
        db.commit()
        pf_id = cur.lastrowid
    return redirect(url_for('portafoglio_dettaglio', pf_id=pf_id))


@app.route('/portafogli/<int:pf_id>')
@login_required
def portafoglio_dettaglio(pf_id):
    anda, andp = uid_and()
    with get_db() as db:
        pf = db.execute(
            f'SELECT * FROM portafogli WHERE id=? {anda}', [pf_id] + andp
        ).fetchone()
        if not pf:
            abort(404)
        clienti = db.execute(
            'SELECT * FROM clienti_portafoglio WHERE portafoglio_id=? ORDER BY created_at',
            (pf_id,)
        ).fetchall()
        aw, ap = uid_where()
        offerte = db.execute(
            f'SELECT id,nome_offerta,commodity,consumo_medio FROM offerte {aw} AND stato="ATTIVA" ORDER BY nome_offerta',
            ap
        ).fetchall() if aw else db.execute(
            'SELECT id,nome_offerta,commodity,consumo_medio FROM offerte WHERE stato="ATTIVA" ORDER BY nome_offerta'
        ).fetchall()
        piani = db.execute(
            f'SELECT id,nome_piano FROM piani_provvigionali {aw} ORDER BY nome_piano', ap
        ).fetchall() if aw else db.execute(
            'SELECT id,nome_piano FROM piani_provvigionali ORDER BY nome_piano'
        ).fetchall()
    clienti_d = [dict(c) for c in clienti]
    n = len(clienti_d)
    tot_lordo  = sum(c['margine_lordo']  or 0 for c in clienti_d)
    tot_netto  = sum(c['margine_netto']  or 0 for c in clienti_d)
    tot_provv  = sum(c['totale_provvigioni'] or 0 for c in clienti_d)
    avg_netto  = round(tot_netto / n, 2) if n else 0
    avg_pct    = round(sum(c['margine_percentuale'] or 0 for c in clienti_d) / n, 1) if n else 0
    return render_template('portafogli.html',
        view='dettaglio',
        pf=dict(pf),
        clienti=clienti_d,
        offerte=[dict(o) for o in offerte],
        piani=[dict(p) for p in piani],
        n_clienti=n,
        tot_lordo=round(tot_lordo,2),
        tot_netto=round(tot_netto,2),
        tot_provv=round(tot_provv,2),
        avg_netto=avg_netto,
        avg_pct=avg_pct,
    )


@app.route('/portafogli/<int:pf_id>/elimina', methods=['POST'])
@login_required
def portafoglio_elimina(pf_id):
    anda, andp = uid_and()
    with get_db() as db:
        db.execute(f'DELETE FROM clienti_portafoglio WHERE portafoglio_id=?', (pf_id,))
        db.execute(f'DELETE FROM portafogli WHERE id=? {anda}', [pf_id] + andp)
        db.commit()
    flash('Portafoglio eliminato.', 'info')
    return redirect(url_for('portafogli'))


@app.route('/portafogli/<int:pf_id>/aggiungi-cliente', methods=['POST'])
@login_required
def portafoglio_aggiungi_cliente(pf_id):
    anda, andp = uid_and()
    with get_db() as db:
        pf = db.execute(f'SELECT id FROM portafogli WHERE id=? {anda}', [pf_id]+andp).fetchone()
    if not pf:
        abort(404)
    f           = request.form
    offerta_id  = int(f['offerta_id'])
    piano_id    = int(f['piano_id'])
    consumo_ov  = float(f['consumo']) if f.get('consumo') else None
    nome_cliente = f.get('nome_cliente', '').strip() or f'Cliente {datetime.now().strftime("%d/%m %H:%M")}'
    note        = f.get('note', '').strip()

    r = _calcola_cliente(offerta_id, piano_id, consumo_ov)
    if not r:
        flash('Offerta o piano non trovati. Verifica i dati.', 'error')
        return redirect(url_for('portafoglio_dettaglio', pf_id=pf_id))

    with get_db() as db:
        db.execute('''INSERT INTO clienti_portafoglio
            (portafoglio_id,user_id,nome_cliente,offerta_id,piano_id,consumo_override,note,
             nome_offerta,nome_fornitore,nome_piano,commodity,
             spread_vendita,spread_acquisto,quota_fissa,costo_gestione_pdp,
             margine_lordo,totale_provvigioni,margine_netto,margine_percentuale)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (pf_id, current_user.id, nome_cliente, offerta_id, piano_id,
             consumo_ov or r['consumo_medio'], note,
             r['nome_offerta'], r['nome_fornitore'], r['nome_piano'], r['commodity'],
             r['spread_vendita'], r['spread_acquisto'], r['quota_fissa'], r['costo_gestione_pdp'],
             r['margine_lordo'], r['totale_provvigioni'], r['margine_netto'], r['margine_percentuale']))
        db.commit()
    flash(f'Cliente "{nome_cliente}" aggiunto al portafoglio.', 'success')
    return redirect(url_for('portafoglio_dettaglio', pf_id=pf_id))


@app.route('/portafogli/<int:pf_id>/elimina-cliente/<int:cid>', methods=['POST'])
@login_required
def portafoglio_elimina_cliente(pf_id, cid):
    with get_db() as db:
        db.execute('DELETE FROM clienti_portafoglio WHERE id=? AND portafoglio_id=?', (cid, pf_id))
        db.commit()
    return ('', 204)


@app.route('/portafogli/<int:pf_id>/clienti/<int:cid>/aggiorna', methods=['POST'])
@login_required
def portafoglio_aggiorna_cliente(pf_id, cid):
    """Aggiorna consumo e note di un cliente e ricalcola i suoi margini."""
    anda, andp = uid_and()
    with get_db() as db:
        pf = db.execute(f'SELECT id FROM portafogli WHERE id=? {anda}', [pf_id]+andp).fetchone()
        if not pf:
            abort(404)
        c = db.execute('SELECT * FROM clienti_portafoglio WHERE id=? AND portafoglio_id=?', (cid, pf_id)).fetchone()
        if not c:
            abort(404)
    consumo_str = request.form.get('consumo', '').strip()
    note = request.form.get('note', c['note'] or '')
    consumo_ov = float(consumo_str) if consumo_str else c['consumo_override']
    # Ricalcola margini con nuovo consumo
    r = _calcola_cliente(c['offerta_id'], c['piano_id'], consumo_ov)
    if r:
        with get_db() as db:
            db.execute('''UPDATE clienti_portafoglio SET
                consumo_override=?, note=?,
                nome_offerta=?,nome_fornitore=?,nome_piano=?,commodity=?,
                spread_vendita=?,spread_acquisto=?,quota_fissa=?,costo_gestione_pdp=?,
                margine_lordo=?,totale_provvigioni=?,margine_netto=?,margine_percentuale=?
                WHERE id=? AND portafoglio_id=?''',
                (consumo_ov, note,
                 r['nome_offerta'],r['nome_fornitore'],r['nome_piano'],r['commodity'],
                 r['spread_vendita'],r['spread_acquisto'],r['quota_fissa'],r['costo_gestione_pdp'],
                 r['margine_lordo'],r['totale_provvigioni'],r['margine_netto'],r['margine_percentuale'],
                 cid, pf_id))
            db.commit()
        flash('Cliente aggiornato e margini ricalcolati.', 'success')
    else:
        with get_db() as db:
            db.execute('UPDATE clienti_portafoglio SET consumo_override=?, note=? WHERE id=? AND portafoglio_id=?',
                       (consumo_ov, note, cid, pf_id))
            db.commit()
        flash('Consumo aggiornato (offerta o piano non disponibili per il ricalcolo).', 'warning')
    return redirect(url_for('portafoglio_dettaglio', pf_id=pf_id))


@app.route('/portafogli/<int:pf_id>/ricalcola', methods=['POST'])
@login_required
def portafoglio_ricalcola(pf_id):
    """Ricalcola i margini di tutti i clienti usando i dati attuali di offerta/fornitore/piano."""
    anda, andp = uid_and()
    with get_db() as db:
        pf = db.execute(f'SELECT id FROM portafogli WHERE id=? {anda}', [pf_id]+andp).fetchone()
        if not pf:
            abort(404)
        clienti = db.execute(
            'SELECT * FROM clienti_portafoglio WHERE portafoglio_id=?', (pf_id,)
        ).fetchall()
    aggiornati = 0
    for c in clienti:
        if c['offerta_id'] and c['piano_id']:
            r = _calcola_cliente(c['offerta_id'], c['piano_id'], c['consumo_override'])
            if r:
                with get_db() as db:
                    db.execute('''UPDATE clienti_portafoglio SET
                        nome_offerta=?,nome_fornitore=?,nome_piano=?,commodity=?,
                        spread_vendita=?,spread_acquisto=?,quota_fissa=?,costo_gestione_pdp=?,
                        margine_lordo=?,totale_provvigioni=?,margine_netto=?,margine_percentuale=?
                        WHERE id=?''',
                        (r['nome_offerta'],r['nome_fornitore'],r['nome_piano'],r['commodity'],
                         r['spread_vendita'],r['spread_acquisto'],r['quota_fissa'],r['costo_gestione_pdp'],
                         r['margine_lordo'],r['totale_provvigioni'],r['margine_netto'],r['margine_percentuale'],
                         c['id']))
                    db.commit()
                aggiornati += 1
    flash(f'Ricalcolo completato: {aggiornati} clienti aggiornati.', 'success')
    return redirect(url_for('portafoglio_dettaglio', pf_id=pf_id))


# ── Template Excel per import ────────────────────────────────────────────────
@app.route('/portafogli/template-import')
@login_required
def portafoglio_template_import():
    aw, ap = uid_where()
    with get_db() as db:
        offerte = db.execute(
            f'SELECT nome_offerta,commodity,consumo_medio FROM offerte {aw} AND stato="ATTIVA" ORDER BY nome_offerta', ap
        ).fetchall() if aw else db.execute(
            'SELECT nome_offerta,commodity,consumo_medio FROM offerte WHERE stato="ATTIVA" ORDER BY nome_offerta'
        ).fetchall()
        piani = db.execute(
            f'SELECT nome_piano FROM piani_provvigionali {aw} ORDER BY nome_piano', ap
        ).fetchall() if aw else db.execute(
            'SELECT nome_piano FROM piani_provvigionali ORDER BY nome_piano'
        ).fetchall()

    wb = Workbook()
    ws = wb.active; ws.title = 'Clienti'
    ws.sheet_view.showGridLines = False

    hfill = PatternFill('solid', fgColor='1a1f2e')
    hfont = Font(bold=True, color='7c9dff', size=10)
    ha    = Alignment(horizontal='center', vertical='center')

    headers = [('Nome Cliente',22),('Offerta (nome esatto)',28),
               ('Piano Provvigionale (nome esatto)',30),('Consumo (lascia vuoto = default offerta)',36),('Note',24)]
    for col,(h,w) in enumerate(headers,1):
        c = ws.cell(row=1,column=col,value=h)
        c.font=hfont; c.fill=hfill; c.alignment=ha
        ws.column_dimensions[chr(64+col)].width=w
    ws.row_dimensions[1].height=22

    # Esempio
    ex_offerta = offerte[0]['nome_offerta'] if offerte else 'NOME_OFFERTA'
    ex_piano   = piani[0]['nome_piano']     if piani   else 'NOME_PIANO'
    ef = PatternFill('solid',fgColor='161b22')
    for i in range(1,6):
        for col in range(1,6):
            c = ws.cell(row=i+1,column=col,value='' if col!=2 else (ex_offerta if i==1 else ''))
            c.font=Font(color='e6edf3',size=10); c.fill=ef
            c.alignment=Alignment(vertical='center')
        ws.cell(row=i+1,column=1,value=f'Cliente {i}').font=Font(color='e6edf3',size=10)
        if i==1:
            ws.cell(row=2,column=2,value=ex_offerta)
            ws.cell(row=2,column=3,value=ex_piano)
        ws.row_dimensions[i+1].height=16

    # Foglio di riferimento con offerte/piani disponibili
    ws2 = wb.create_sheet('Riferimento (non modificare)')
    ws2.cell(row=1,column=1,value='OFFERTE DISPONIBILI').font=Font(bold=True,color='7c9dff',size=10)
    ws2.cell(row=1,column=4,value='PIANI DISPONIBILI').font=Font(bold=True,color='7c9dff',size=10)
    for i,o in enumerate(offerte,2):
        ws2.cell(row=i,column=1,value=o['nome_offerta']).font=Font(color='e6edf3',size=10)
        ws2.cell(row=i,column=2,value=o['commodity']).font=Font(color='8b949e',size=9)
        ws2.cell(row=i,column=3,value=f'Default: {o["consumo_medio"]}').font=Font(color='8b949e',size=9)
    for i,p in enumerate(piani,2):
        ws2.cell(row=i,column=4,value=p['nome_piano']).font=Font(color='e6edf3',size=10)
    ws2.column_dimensions['A'].width=28; ws2.column_dimensions['D'].width=28

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='template_portafoglio.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── Import clienti da file ────────────────────────────────────────────────────
@app.route('/portafogli/<int:pf_id>/import', methods=['POST'])
@login_required
def portafoglio_import(pf_id):
    anda, andp = uid_and()
    with get_db() as db:
        pf = db.execute(f'SELECT id FROM portafogli WHERE id=? {anda}', [pf_id]+andp).fetchone()
    if not pf:
        abort(404)

    file = request.files.get('file')
    if not file or not file.filename:
        flash('Nessun file selezionato.', 'warning')
        return redirect(url_for('portafoglio_dettaglio', pf_id=pf_id))

    fname = file.filename.lower()
    rows  = []

    try:
        if fname.endswith('.csv'):
            content = file.read().decode('utf-8-sig')
            reader  = csv.DictReader(io.StringIO(content))
            for r in reader:
                rows.append({k.strip(): v.strip() for k,v in r.items()})
        elif fname.endswith('.xlsx'):
            wb2 = load_workbook(io.BytesIO(file.read()), read_only=True, data_only=True)
            ws2 = wb2.active
            headers2 = [str(c.value).strip() if c.value else '' for c in next(ws2.iter_rows(min_row=1,max_row=1))]
            for row in ws2.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                rows.append({headers2[i]: (str(v).strip() if v is not None else '') for i,v in enumerate(row)})
            wb2.close()
        else:
            flash('Formato non supportato. Usa .xlsx o .csv.', 'error')
            return redirect(url_for('portafoglio_dettaglio', pf_id=pf_id))
    except Exception as e:
        flash(f'Errore nella lettura del file: {e}', 'error')
        return redirect(url_for('portafoglio_dettaglio', pf_id=pf_id))

    # Normalizza nomi colonne (flessibile)
    def _col(r, *keys):
        for k in keys:
            for rk in r:
                if k.lower() in rk.lower():
                    return r[rk]
        return ''

    aw2, ap2 = uid_where()
    with get_db() as db:
        offerte_db = {o['nome_offerta'].lower(): dict(o)
                      for o in db.execute(f'SELECT * FROM offerte {aw2} AND stato="ATTIVA"', ap2).fetchall()} \
                     if aw2 else \
                     {o['nome_offerta'].lower(): dict(o)
                      for o in db.execute('SELECT * FROM offerte WHERE stato="ATTIVA"').fetchall()}
        piani_db   = {p['nome_piano'].lower(): dict(p)
                      for p in db.execute(f'SELECT * FROM piani_provvigionali {aw2}', ap2).fetchall()} \
                     if aw2 else \
                     {p['nome_piano'].lower(): dict(p)
                      for p in db.execute('SELECT * FROM piani_provvigionali').fetchall()}

    ok = 0; skipped = 0; errors = []
    for i, row in enumerate(rows, 1):
        nome_cliente = _col(row,'nome cliente','cliente','name') or f'Cliente {i}'
        nome_offerta = _col(row,'offerta')
        nome_piano   = _col(row,'piano')
        consumo_str  = _col(row,'consumo')
        note         = _col(row,'note','notes')

        offerta_d = offerte_db.get(nome_offerta.lower())
        piano_d   = piani_db.get(nome_piano.lower())

        if not offerta_d:
            errors.append(f'Riga {i} ("{nome_cliente}"): offerta "{nome_offerta}" non trovata')
            skipped += 1; continue
        if not piano_d:
            errors.append(f'Riga {i} ("{nome_cliente}"): piano "{nome_piano}" non trovato')
            skipped += 1; continue

        try:
            consumo_ov = float(consumo_str) if consumo_str else None
        except ValueError:
            consumo_ov = None

        fornitore_d = _get_fornitore_for_api(None, offerta_d)
        if not fornitore_d:
            errors.append(f'Riga {i} ("{nome_cliente}"): nessun fornitore per {offerta_d["commodity"]}')
            skipped += 1; continue

        # Controllo doppioni: skip se nome_cliente già presente nel portafoglio
        with get_db() as db:
            dup = db.execute(
                'SELECT id FROM clienti_portafoglio WHERE portafoglio_id=? AND LOWER(nome_cliente)=LOWER(?)',
                [pf_id, nome_cliente]
            ).fetchone()
        if dup:
            errors.append(f'Riga {i} ("{nome_cliente}"): già presente — saltato')
            skipped += 1
            continue

        r = calcola_simulazione(offerta_d, fornitore_d, piano_d, consumo_ov)
        with get_db() as db:
            db.execute('''INSERT INTO clienti_portafoglio
                (portafoglio_id,user_id,nome_cliente,offerta_id,piano_id,consumo_override,note,
                 nome_offerta,nome_fornitore,nome_piano,commodity,
                 spread_vendita,spread_acquisto,quota_fissa,costo_gestione_pdp,
                 margine_lordo,totale_provvigioni,margine_netto,margine_percentuale)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                (pf_id, current_user.id, nome_cliente,
                 offerta_d['id'], piano_d['id'],
                 consumo_ov or r['consumo_medio'], note,
                 offerta_d['nome_offerta'], fornitore_d.get('nome','Media automatica'),
                 piano_d['nome_piano'], offerta_d['commodity'],
                 r['spread_vendita'], r['spread_acquisto'], r['quota_fissa'], r['costo_gestione_pdp'],
                 r['margine_lordo'], r['totale_provvigioni'], r['margine_netto'], r['margine_percentuale']))
            db.commit()
        ok += 1

    msg = f'Import completato: {ok} clienti aggiunti'
    if skipped:
        msg += f', {skipped} righe saltate'
    flash(msg + ('.  Errori: ' + ' | '.join(errors[:3]) if errors else '.'), 'success' if ok else 'warning')
    return redirect(url_for('portafoglio_dettaglio', pf_id=pf_id))


# ── Export Excel report portafoglio ──────────────────────────────────────────
@app.route('/portafogli/<int:pf_id>/export')
@login_required
def portafoglio_export(pf_id):
    anda, andp = uid_and()
    with get_db() as db:
        pf = db.execute(f'SELECT * FROM portafogli WHERE id=? {anda}', [pf_id]+andp).fetchone()
        if not pf:
            abort(404)
        clienti = db.execute(
            'SELECT * FROM clienti_portafoglio WHERE portafoglio_id=? ORDER BY margine_netto DESC',
            (pf_id,)
        ).fetchall()
    pf_d = dict(pf); clienti_d = [dict(c) for c in clienti]
    n = len(clienti_d)
    tot_lordo = sum(c['margine_lordo'] or 0 for c in clienti_d)
    tot_netto = sum(c['margine_netto'] or 0 for c in clienti_d)
    tot_provv = sum(c['totale_provvigioni'] or 0 for c in clienti_d)
    avg_pct   = sum(c['margine_percentuale'] or 0 for c in clienti_d) / n if n else 0

    BG0='0d1117'; BG1='161b22'; BG2='1a1f2e'
    FG='e6edf3'; FGA='7c9dff'; FGM='8b949e'
    GRN='3fb950'; RED='f85149'

    wb = Workbook()

    # ── Sheet 1: Riepilogo ──────────────────────────────────────────────────
    ws = wb.active; ws.title = 'Riepilogo'
    ws.sheet_view.showGridLines = False
    for col,w in [('A',28),('B',16),('C',3),('D',24),('E',16)]:
        ws.column_dimensions[col].width = w

    ws.merge_cells('A1:E1')
    t = ws.cell(row=1,column=1,value=f'PORTAFOGLIO: {pf_d["nome"].upper()}')
    t.font=Font(color=FGA,bold=True,size=14); t.fill=PatternFill('solid',fgColor=BG0)
    t.alignment=Alignment(horizontal='left',vertical='center'); ws.row_dimensions[1].height=30

    ws.merge_cells('A2:E2')
    sub=ws.cell(row=2,column=1,
        value=f'Generato il {datetime.now().strftime("%d/%m/%Y %H:%M")}  •  {n} clienti  •  {pf_d.get("descrizione") or ""}')
    sub.font=Font(color=FGM,size=9,italic=True); sub.fill=PatternFill('solid',fgColor=BG0)
    sub.alignment=Alignment(horizontal='left',vertical='center'); ws.row_dimensions[2].height=14
    ws.row_dimensions[3].height=8

    def kv(row,label,val,vc=None,fmt=None,bold=False):
        c1=ws.cell(row=row,column=1,value=label)
        c1.font=Font(color=FGM,size=9); c1.alignment=Alignment(vertical='center')
        c2=ws.cell(row=row,column=2,value=val)
        c2.font=Font(color=vc or FG,bold=bold,size=10)
        c2.alignment=Alignment(horizontal='right',vertical='center')
        c2.fill=PatternFill('solid',fgColor=BG1 if row%2==0 else BG0)
        if fmt: c2.number_format=fmt
        ws.row_dimensions[row].height=16

    def sec(row,label,ncol=2):
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=ncol)
        c=ws.cell(row=row,column=1,value=f'  {label}')
        c.font=Font(color=FGA,bold=True,size=9); c.fill=PatternFill('solid',fgColor=BG0)
        c.alignment=Alignment(horizontal='left',vertical='center'); ws.row_dimensions[row].height=15

    r=4; sec(r,'RIEPILOGO PORTAFOGLIO'); r+=1
    kv(r,'Numero clienti',n,bold=True); r+=1
    kv(r,'Margine lordo totale',tot_lordo,fmt='€#,##0.00',vc=FGA,bold=True); r+=1
    kv(r,'Provvigioni totali',tot_provv,fmt='€#,##0.00',vc=RED); r+=1
    kv(r,'Margine netto totale',tot_netto,fmt='€#,##0.00',vc=GRN if tot_netto>=0 else RED,bold=True); r+=1
    kv(r,'Margine % medio',f'{avg_pct:.1f}%',vc=GRN if avg_pct>=0 else RED); r+=1
    kv(r,'Margine netto medio/cliente',round(tot_netto/n,2) if n else 0,fmt='€#,##0.00'); r+=1

    ws.row_dimensions[r].height=8; r+=1

    # Breakdown commodity
    luce = [c for c in clienti_d if (c.get('commodity') or '').upper()=='LUCE']
    gas  = [c for c in clienti_d if (c.get('commodity') or '').upper()=='GAS']
    if luce or gas:
        sec(r,'BREAKDOWN COMMODITY'); r+=1
        if luce:
            kv(r,f'🔆 LUCE — {len(luce)} clienti',
               sum(c['margine_netto'] or 0 for c in luce),fmt='€#,##0.00',vc='f0883e'); r+=1
        if gas:
            kv(r,f'🔥 GAS — {len(gas)} clienti',
               sum(c['margine_netto'] or 0 for c in gas),fmt='€#,##0.00',vc='58a6ff'); r+=1
        ws.row_dimensions[r].height=8; r+=1

    # Breakdown per offerta
    from collections import defaultdict
    by_offerta = defaultdict(list)
    for c in clienti_d:
        by_offerta[c.get('nome_offerta') or '—'].append(c)
    if len(by_offerta) > 1:
        sec(r,'BREAKDOWN PER OFFERTA'); r+=1
        for off_nome, grp in sorted(by_offerta.items()):
            tot = sum(c['margine_netto'] or 0 for c in grp)
            kv(r, f'{off_nome} ({len(grp)} clienti)', tot, fmt='€#,##0.00',
               vc=GRN if tot>=0 else RED); r+=1
        ws.row_dimensions[r].height=8; r+=1

    # ── Bar chart netto per commodity (col D) ──
    if luce and gas:
        ws.cell(row=4,column=4,value='Commodity').font=Font(color=FGM,size=9)
        ws.cell(row=4,column=5,value='M. Netto (€)').font=Font(color=FGM,size=9)
        ws.cell(row=5,column=4,value='LUCE')
        ws.cell(row=5,column=5,value=sum(c['margine_netto'] or 0 for c in luce))
        ws.cell(row=6,column=4,value='GAS')
        ws.cell(row=6,column=5,value=sum(c['margine_netto'] or 0 for c in gas))
        bc=BarChart(); bc.type='col'; bc.grouping='clustered'
        bc.title=None; bc.legend=None; bc.style=10; bc.width=14; bc.height=10
        bc.add_data(Reference(ws,min_col=5,max_col=5,min_row=4,max_row=6),titles_from_data=True)
        bc.set_categories(Reference(ws,min_col=4,max_col=4,min_row=5,max_row=6))
        bc.series[0].graphicalProperties.solidFill='4f6ef7'
        bc.anchor='D9'; ws.add_chart(bc)

    # ── Sheet 2: Lista clienti ──────────────────────────────────────────────
    ws2 = wb.create_sheet('Clienti'); ws2.sheet_view.showGridLines=False
    col2=[('Nome Cliente',22),('Offerta',22),('Fornitore',18),('Piano',18),
          ('Commodity',12),('Consumo',14),('M. Lordo (€)',14),('Provv. Tot. (€)',14),
          ('M. Netto (€)',14),('M. %',10),('Note',20)]
    for i,(h,w) in enumerate(col2,1):
        ws2.column_dimensions[chr(64+i)].width=w
        c=ws2.cell(row=1,column=i,value=h)
        c.font=Font(bold=True,color=FGA,size=9)
        c.fill=PatternFill('solid',fgColor=BG2)
        c.alignment=Alignment(horizontal='center',vertical='center')
        c.border=Border(bottom=Side(style='thin',color='30363d'))
    ws2.row_dimensions[1].height=18

    for ri,cl in enumerate(clienti_d,2):
        alt=ri%2==0; bg=BG1 if alt else BG0
        mnc=GRN if (cl['margine_netto'] or 0)>=0 else RED
        vals=[cl.get('nome_cliente') or '—', cl.get('nome_offerta') or '—',
              cl.get('nome_fornitore') or '—', cl.get('nome_piano') or '—',
              cl.get('commodity') or '—', cl.get('consumo_override'),
              cl.get('margine_lordo'), cl.get('totale_provvigioni'),
              cl.get('margine_netto'), f'{cl["margine_percentuale"]:.1f}%' if cl.get("margine_percentuale") is not None else '—',
              cl.get('note') or '—']
        fmts=[None,None,None,None,None,'#,##0.0000','€#,##0.00','€#,##0.00','€#,##0.00',None,None]
        cols_vc=[None,None,None,None,None,None,FGA,RED,mnc,mnc,None]
        bolds=[False,False,False,False,False,False,False,False,True,False,False]
        for ci,(v,fmt,vc,bold) in enumerate(zip(vals,fmts,cols_vc,bolds),1):
            c=ws2.cell(row=ri,column=ci,value=v)
            c.font=Font(color=vc or FG,bold=bold,size=10)
            c.fill=PatternFill('solid',fgColor=bg)
            c.alignment=Alignment(vertical='center')
            if fmt: c.number_format=fmt
        ws2.row_dimensions[ri].height=16

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    safe=pf_d['nome'].replace(' ','_')[:20]
    fname=f'portafoglio_{safe}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(buf,as_attachment=True,download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ══════════════════════════════════════════════════════════════════════════════
# CONFRONTO SCENARI
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/export/simulazione/<int:sim_id>/pdf')
@login_required
def export_simulazione_pdf(sim_id):
    # ── Carica dati ──────────────────────────────────────────────────────────
    anda, andp = uid_and()
    with get_db() as db:
        sim = db.execute(f'SELECT * FROM simulazioni WHERE id=? {anda}', [sim_id]+andp).fetchone()
    if not sim:
        abort(404)
    sim = dict(sim)

    offerta_v   = {'spread': sim['spread_vendita'], 'quota_fissa': sim['quota_fissa'],
                   'consumo_medio': sim['consumo_medio']}
    fornitore_v = {'spread_acquisto': sim['spread_acquisto'],
                   'costo_gestione_pdp': sim['costo_gestione_pdp'], 'nome': sim['nome_fornitore']}
    piano_d = None
    if sim.get('nome_piano'):
        aw2, ap2 = uid_where()
        sep = 'AND' if aw2 else 'WHERE'
        with get_db() as db:
            pr = db.execute(f'SELECT * FROM piani_provvigionali {aw2} {sep} nome_piano=?',
                            ap2+[sim['nome_piano']]).fetchone()
        if pr: piano_d = dict(pr)

    base = sim['consumo_medio']
    sensitivity = []
    for pct in [25, 50, 75, 100, 125, 150, 175]:
        c = round(base * pct / 100, 4)
        if piano_d:
            r = calcola_simulazione(offerta_v, fornitore_v, piano_d, c)
        else:
            sn = sim['spread_vendita']-sim['spread_acquisto']
            msa = round(sn*c,2); qfa = round((sim['quota_fissa']-sim['costo_gestione_pdp'])*12,2)
            ml=round(msa+qfa,2); pp=sim['totale_provvigioni']; mn=round(ml-pp,2)
            r=dict(consumo_medio=c,margine_spread_annuo=msa,margine_qf_annuo=qfa,
                   margine_lordo=ml,totale_provvigioni=pp,margine_netto=mn,
                   margine_percentuale=round(mn/ml*100,2) if ml else 0)
        sensitivity.append({'pct':pct,'consumo':c,**r})

    breakeven_val = None
    if piano_d:
        sn=sim['spread_vendita']-sim['spread_acquisto']
        rct=(piano_d['ricorrente_consumo_agente']+piano_d['ricorrente_consumo_sub_agente']
             +piano_d['ricorrente_consumo_area_manager'])
        gf=(piano_d['gettone_agente']+piano_d['gettone_sub_agente']
            +(piano_d['ricorrente_mese_agente']+piano_d['ricorrente_mese_sub_agente']
              +piano_d['ricorrente_mese_area_manager'])*12)
        qfn=(sim['quota_fissa']-sim['costo_gestione_pdp'])*12
        coef=sn-rct
        if coef>0: breakeven_val=round(max(0,(gf-qfn)/coef),4)

    unit = 'MWh' if sim.get('commodity')=='LUCE' else 'smc'

    # ── Palette colori per PDF (light-on-dark adattato a stampa) ─────────────
    C_BG     = colors.HexColor('#0d1117')
    C_BG2    = colors.HexColor('#161b22')
    C_BG3    = colors.HexColor('#1a1f2e')
    C_ACCENT = colors.HexColor('#4f6ef7')
    C_PURPLE = colors.HexColor('#bc8cff')
    C_GREEN  = colors.HexColor('#3fb950')
    C_RED    = colors.HexColor('#f85149')
    C_ORANGE = colors.HexColor('#f0883e')
    C_TEXT   = colors.HexColor('#e6edf3')
    C_MUTED  = colors.HexColor('#8b949e')
    C_BORDER = colors.HexColor('#30363d')
    C_WHITE  = colors.white

    def style(name, **kw):
        base_kw = dict(fontName='Helvetica', fontSize=10, textColor=C_TEXT,
                       backColor=C_BG, leading=14)
        base_kw.update(kw)
        return ParagraphStyle(name, **base_kw)

    S_TITLE   = style('title',  fontName='Helvetica-Bold', fontSize=18, textColor=C_ACCENT, leading=22)
    S_SUB     = style('sub',    fontSize=9, textColor=C_MUTED, leading=12)
    S_SEC     = style('sec',    fontName='Helvetica-Bold', fontSize=8, textColor=C_ACCENT,
                      spaceAfter=4, spaceBefore=10, leading=10)
    S_BODY    = style('body',   fontSize=9, textColor=C_TEXT, leading=13)
    S_BOLD    = style('bold',   fontName='Helvetica-Bold', fontSize=10, textColor=C_TEXT, leading=14)
    S_GREEN   = style('green',  fontName='Helvetica-Bold', fontSize=12, textColor=C_GREEN, leading=16)
    S_RED_B   = style('redb',   fontName='Helvetica-Bold', fontSize=12, textColor=C_RED,   leading=16)

    def ts_base(extra=None):
        cmds = [
            ('BACKGROUND',    (0,0),(-1,0),  C_BG3),
            ('BACKGROUND',    (0,1),(-1,-1), C_BG2),
            ('ROWBACKGROUNDS',(0,1),(-1,-1), [C_BG2, colors.HexColor('#0f1419')]),
            ('TEXTCOLOR',     (0,0),(-1,-1), C_TEXT),
            ('FONTNAME',      (0,0),(-1,0),  'Helvetica-Bold'),
            ('FONTSIZE',       (0,0),(-1,0),  8),
            ('FONTSIZE',       (0,1),(-1,-1), 9),
            ('GRID',          (0,0),(-1,-1), 0.3, C_BORDER),
            ('ALIGN',         (0,0),(0,-1),  'LEFT'),
            ('ALIGN',         (1,0),(-1,-1), 'RIGHT'),
            ('TOPPADDING',    (0,0),(-1,-1), 5),
            ('BOTTOMPADDING', (0,0),(-1,-1), 5),
            ('LEFTPADDING',   (0,0),(0,-1),  8),
            ('RIGHTPADDING',  (1,0),(-1,-1), 8),
        ]
        if extra: cmds.extend(extra)
        return TableStyle(cmds)

    def euro(v): return f'€ {v:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
    def sign_euro(v): return ('+' if v>0 else '')+euro(v)

    buf  = io.BytesIO()
    page_w, page_h = A4

    def on_page(canvas, doc):
        """Header e footer su ogni pagina."""
        canvas.saveState()
        # Sfondo pagina scuro
        canvas.setFillColor(C_BG)
        canvas.rect(0, 0, page_w, page_h, fill=1, stroke=0)
        # Striscia accent in cima
        canvas.setFillColor(C_ACCENT)
        canvas.rect(0, page_h-4, page_w, 4, fill=1, stroke=0)
        # Footer
        canvas.setFillColor(C_MUTED)
        canvas.setFont('Helvetica', 7)
        canvas.drawString(2*cm, 1.2*cm,
            f'Energia Simulator  •  Simulazione #{sim_id}  •  {datetime.now().strftime("%d/%m/%Y")}')
        canvas.drawRightString(page_w-2*cm, 1.2*cm, f'Pag. {doc.page}')
        canvas.restoreState()

    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm, topMargin=2.5*cm, bottomMargin=2.5*cm)

    story = []
    data_str = (sim.get('created_at') or '')[:16].replace('T', ' ')

    # ── Titolo ────────────────────────────────────────────────────────────────
    story.append(Paragraph(f'ANALISI SIMULAZIONE', S_TITLE))
    story.append(Paragraph(sim.get('nome_offerta') or '', style('off', fontName='Helvetica-Bold',
        fontSize=14, textColor=C_PURPLE, leading=18)))
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph(
        f'Simulazione del {data_str}  &nbsp;|&nbsp;  Piano: {sim.get("nome_piano") or "—"}  &nbsp;|&nbsp;  {sim.get("commodity") or ""}  {sim.get("tipo_consumo") or ""}',
        S_SUB))
    story.append(HRFlowable(width='100%', thickness=0.5, color=C_BORDER, spaceAfter=10, spaceBefore=6))

    # ── KPI griglia 2×2 ───────────────────────────────────────────────────────
    mn = sim.get('margine_netto') or 0
    ml = sim.get('margine_lordo') or 0
    pp = sim.get('totale_provvigioni') or 0
    pct= sim.get('margine_percentuale') or 0
    mn_color = C_GREEN if mn >= 0 else C_RED

    kpi_data = [
        [Paragraph('MARGINE LORDO', style('kl',fontSize=8,textColor=C_MUTED,backColor=C_BG3)),
         Paragraph('PROVVIGIONI TOTALI', style('kl',fontSize=8,textColor=C_MUTED,backColor=C_BG3)),
         Paragraph('MARGINE NETTO', style('kl',fontSize=8,textColor=C_MUTED,backColor=C_BG3)),
         Paragraph('MARGINE %', style('kl',fontSize=8,textColor=C_MUTED,backColor=C_BG3))],
        [Paragraph(euro(ml), style('kv',fontName='Helvetica-Bold',fontSize=14,textColor=C_ACCENT,backColor=C_BG3,leading=18)),
         Paragraph(euro(pp), style('kv',fontName='Helvetica-Bold',fontSize=14,textColor=C_RED,backColor=C_BG3,leading=18)),
         Paragraph(euro(mn), style('kv',fontName='Helvetica-Bold',fontSize=14,textColor=mn_color,backColor=C_BG3,leading=18)),
         Paragraph(f'{pct:.1f}%', style('kv',fontName='Helvetica-Bold',fontSize=14,textColor=mn_color,backColor=C_BG3,leading=18))],
    ]
    kpi_tbl = Table(kpi_data, colWidths=[(page_w-4*cm)/4]*4, rowHeights=[16, 28])
    kpi_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0),(-1,-1), C_BG3),
        ('GRID',       (0,0),(-1,-1), 0.3, C_BORDER),
        ('TOPPADDING', (0,0),(-1,-1), 6),
        ('BOTTOMPADDING',(0,0),(-1,-1),6),
        ('LEFTPADDING',(0,0),(-1,-1),10),
        ('ROUNDEDCORNERS',[4,4,4,4]),
    ]))
    story.append(kpi_tbl)
    story.append(Spacer(1, 0.5*cm))

    # ── Dati simulazione + Parametri tecnici (2 colonne) ─────────────────────
    sn_val = round((sim['spread_vendita'] or 0)-(sim['spread_acquisto'] or 0), 4)

    left_rows = [
        [Paragraph('DATI SIMULAZIONE', style('sh',fontName='Helvetica-Bold',fontSize=8,textColor=C_ACCENT,backColor=C_BG3)), ''],
        ['Offerta (CTE)', sim.get('nome_offerta') or '—'],
        ['Fornitore',     sim.get('nome_fornitore') or '—'],
        ['Piano',         sim.get('nome_piano') or '—'],
        ['Commodity',     f"{sim.get('commodity') or '—'} — {sim.get('tipo_consumo') or '—'}"],
        [f'Consumo ({unit})', f"{sim.get('consumo_medio') or 0:,.4f}"],
        ['Data',          data_str],
        ['Note',          sim.get('note') or '—'],
    ]
    right_rows = [
        [Paragraph('PARAMETRI TECNICI', style('sh',fontName='Helvetica-Bold',fontSize=8,textColor=C_ACCENT,backColor=C_BG3)), ''],
        ['Spread vendita',  f"€ {sim.get('spread_vendita') or 0:.4f}/unità"],
        ['Spread acquisto', f"€ {sim.get('spread_acquisto') or 0:.4f}/unità"],
        ['Spread netto',    Paragraph(f"€ {sn_val:.4f}/unità",
            style('sn',fontName='Helvetica-Bold',fontSize=9,textColor=C_GREEN if sn_val>=0 else C_RED,backColor=C_BG2))],
        ['QF vendita',      f"€ {sim.get('quota_fissa') or 0:.2f}/mese"],
        ['Costo PDP',       f"€ {sim.get('costo_gestione_pdp') or 0:.2f}/mese"],
        ['Breakeven',       (f"{breakeven_val:,.2f} {unit}" if breakeven_val is not None else 'N/D')],
        ['', ''],
    ]

    def info_tbl(rows, col_w):
        tbl = Table(rows, colWidths=col_w)
        cmds = [
            ('BACKGROUND', (0,0),(-1,0),  C_BG3),
            ('SPAN',       (0,0),(-1,0)),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[C_BG2, colors.HexColor('#0f1419')]),
            ('TEXTCOLOR',  (0,0),(-1,-1), C_TEXT),
            ('FONTNAME',   (0,1),(0,-1),  'Helvetica'),
            ('FONTSIZE',   (0,0),(-1,-1), 8),
            ('FONTNAME',   (1,1),(-1,-1), 'Helvetica-Bold'),
            ('GRID',       (0,0),(-1,-1), 0.3, C_BORDER),
            ('TOPPADDING', (0,0),(-1,-1), 4),
            ('BOTTOMPADDING',(0,0),(-1,-1),4),
            ('LEFTPADDING',(0,0),(-1,-1), 8),
            ('ALIGN',      (1,0),(-1,-1), 'RIGHT'),
            ('RIGHTPADDING',(1,0),(-1,-1),8),
        ]
        tbl.setStyle(TableStyle(cmds))
        return tbl

    half = (page_w-4*cm-0.4*cm)/2
    side_tbl = Table([[info_tbl(left_rows,[half*0.55,half*0.45]),
                       Spacer(0.4*cm,1),
                       info_tbl(right_rows,[half*0.55,half*0.45])]], colWidths=[half,0.4*cm,half])
    side_tbl.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'TOP'),('LEFTPADDING',(0,0),(-1,-1),0),('RIGHTPADDING',(0,0),(-1,-1),0)]))
    story.append(side_tbl)
    story.append(Spacer(1, 0.5*cm))

    # ── Grafico a barre: composizione margini ─────────────────────────────────
    story.append(Paragraph('COMPOSIZIONE MARGINI', S_SEC))
    bar_labels = ['M. Spread', 'M. QF', 'Provv. Ag.', 'Provv. Sub', 'Provv. AM', 'M. Netto']
    bar_values = [sim.get('margine_spread_annuo') or 0, sim.get('margine_qf_annuo') or 0,
                  -(sim.get('provvigione_agente') or 0), -(sim.get('provvigione_sub_agente') or 0),
                  -(sim.get('provvigione_area_manager') or 0), sim.get('margine_netto') or 0]
    bar_colors = [colors.HexColor('#4f6ef7'), colors.HexColor('#7c9dff'),
                  colors.HexColor('#f85149'), colors.HexColor('#d29922'),
                  colors.HexColor('#bc8cff'), C_GREEN if mn>=0 else C_RED]

    chart_w, chart_h = 14*cm, 6*cm
    d = Drawing(chart_w, chart_h)
    d.add(Rect(0,0,chart_w,chart_h,fillColor=C_BG2,strokeColor=C_BORDER,strokeWidth=0.3))
    bc = VerticalBarChart()
    bc.x = 1.2*cm; bc.y = 0.8*cm
    bc.width  = chart_w - 2.0*cm
    bc.height = chart_h - 1.2*cm
    bc.data = [bar_values]
    bc.categoryAxis.categoryNames = bar_labels
    bc.categoryAxis.labels.fontSize = 7
    bc.categoryAxis.labels.fillColor = C_MUTED
    bc.categoryAxis.strokeColor = C_BORDER
    bc.valueAxis.labels.fontSize = 7
    bc.valueAxis.labels.fillColor = C_MUTED
    bc.valueAxis.strokeColor = C_BORDER
    bc.valueAxis.gridStrokeColor = C_BORDER
    bc.valueAxis.gridStrokeWidth = 0.3
    bc.bars[0].fillColor = colors.HexColor('#4f6ef7')
    for i,c_col in enumerate(bar_colors):
        bc.bars[0,i].fillColor = c_col
    bc.groupSpacing = 8
    bc.barSpacing = 2
    bc.bars.strokeWidth = 0
    d.add(bc)
    story.append(KeepTogether([d, Spacer(1, 0.4*cm)]))

    # ── Tabella sensitivity ────────────────────────────────────────────────────
    story.append(Paragraph('ANALISI DI SENSITIVITÁ', S_SEC))
    if not piano_d:
        story.append(Paragraph(
            '⚠  Piano non trovato nel DB: le provvigioni sono stimate al valore base della simulazione.',
            style('warn', fontSize=8, textColor=C_ORANGE, leading=11)))
        story.append(Spacer(1, 0.2*cm))

    sens_header = [Paragraph(h, style('th',fontName='Helvetica-Bold',fontSize=8,
                                       textColor=C_ACCENT,backColor=C_BG3))
                   for h in ['% Base', f'Consumo ({unit})', 'M. Lordo', 'Provv. Tot.', 'M. Netto', 'M. %']]
    sens_rows = [sens_header]
    for row in sensitivity:
        is_base = row['pct'] == 100
        mn_r = row.get('margine_netto') or 0
        mn_c = C_GREEN if mn_r >= 0 else C_RED
        bg   = colors.HexColor('#1a2f3a') if is_base else None
        def sp(txt, color=C_TEXT, bold=False, bg_c=bg):
            fs = 9 if not is_base else 9
            fn = 'Helvetica-Bold' if (bold or is_base) else 'Helvetica'
            return Paragraph(str(txt), style(f'sc{id(txt)}',fontName=fn,fontSize=fs,
                                             textColor=color,backColor=bg_c or C_BG2,leading=12))
        sens_rows.append([
            sp(f"{row['pct']}%",      color=C_ACCENT if is_base else C_MUTED),
            sp(f"{row['consumo']:,.4f}"),
            sp(euro(row.get('margine_lordo') or 0)),
            sp(euro(row.get('totale_provvigioni') or 0), color=C_RED),
            sp(euro(mn_r), color=mn_c, bold=True),
            sp(f"{(row.get('margine_percentuale') or 0):.1f}%", color=mn_c),
        ])

    col_w = (page_w-4*cm)/6
    sens_tbl = Table(sens_rows, colWidths=[col_w]*6)
    sens_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0,0),(-1,0), C_BG3),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[C_BG2, colors.HexColor('#0f1419')]),
        ('GRID',          (0,0),(-1,-1),0.3, C_BORDER),
        ('ALIGN',         (0,0),(-1,-1),'RIGHT'),
        ('ALIGN',         (0,0),(0,-1), 'CENTER'),
        ('TOPPADDING',    (0,0),(-1,-1),5),
        ('BOTTOMPADDING', (0,0),(-1,-1),5),
        ('LEFTPADDING',   (0,0),(-1,-1),6),
        ('RIGHTPADDING',  (0,0),(-1,-1),6),
    ]))
    story.append(sens_tbl)
    story.append(Spacer(1, 0.4*cm))

    # ── Grafico sensitivity ────────────────────────────────────────────────────
    sd_w, sd_h = 14*cm, 5.5*cm
    sd = Drawing(sd_w, sd_h)
    sd.add(Rect(0,0,sd_w,sd_h,fillColor=C_BG2,strokeColor=C_BORDER,strokeWidth=0.3))
    lc = HorizontalLineChart()
    lc.x = 1.5*cm; lc.y = 0.8*cm
    lc.width  = sd_w - 2.2*cm
    lc.height = sd_h - 1.4*cm
    lc.data = [
        [s.get('margine_lordo') or 0    for s in sensitivity],
        [s.get('totale_provvigioni') or 0 for s in sensitivity],
        [s.get('margine_netto') or 0    for s in sensitivity],
    ]
    lc.categoryAxis.categoryNames = [f"{s['pct']}%" for s in sensitivity]
    lc.categoryAxis.labels.fontSize = 7
    lc.categoryAxis.labels.fillColor = C_MUTED
    lc.categoryAxis.strokeColor = C_BORDER
    lc.valueAxis.labels.fontSize = 7
    lc.valueAxis.labels.fillColor = C_MUTED
    lc.valueAxis.strokeColor = C_BORDER
    lc.valueAxis.gridStrokeColor = C_BORDER
    lc.valueAxis.gridStrokeWidth = 0.3
    lc.lines[0].strokeColor = colors.HexColor('#4f6ef7')
    lc.lines[0].strokeWidth = 1.5
    lc.lines[1].strokeColor = C_RED
    lc.lines[1].strokeWidth = 1.2
    lc.lines[1].strokeDashArray = [4, 2]
    lc.lines[2].strokeColor = C_GREEN
    lc.lines[2].strokeWidth = 2
    sd.add(lc)
    # Legenda manuale
    for i,(lbl,col) in enumerate([('M. Lordo','#4f6ef7'),('Provvigioni','#f85149'),('M. Netto','#3fb950')]):
        x0 = 1.5*cm + i*4.0*cm
        sd.add(Rect(x0, 0.2*cm, 0.6*cm, 0.2*cm, fillColor=colors.HexColor(col), strokeWidth=0))
        sd.add(String(x0+0.7*cm, 0.18*cm, lbl, fontSize=7, fillColor=C_MUTED))
    story.append(KeepTogether([sd]))

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page)
    buf.seek(0)
    offerta_safe = (sim.get('nome_offerta') or 'sim').replace(' ', '_')[:20]
    fname = f'report_{sim_id}_{offerta_safe}_{datetime.now().strftime("%Y%m%d")}.pdf'
    return send_file(buf, as_attachment=True, download_name=fname, mimetype='application/pdf')


@app.route('/export/confronto')
@login_required
def export_confronto():
    sid_a = request.args.get('a', type=int)
    sid_b = request.args.get('b', type=int)
    if not sid_a or not sid_b:
        abort(400)
    anda, andp = uid_and()
    with get_db() as db:
        sa = db.execute(f'SELECT * FROM simulazioni WHERE id=? {anda}', [sid_a]+andp).fetchone()
        sb = db.execute(f'SELECT * FROM simulazioni WHERE id=? {anda}', [sid_b]+andp).fetchone()
    if not sa or not sb:
        abort(404)
    sa, sb = dict(sa), dict(sb)

    BG0='0d1117'; BG1='161b22'; BG2='1a1f2e'
    FG='e6edf3'; FGA='7c9dff'; FGM='8b949e'; GRN='3fb950'; RED='f85149'; PRP='bc8cff'

    wb = Workbook(); ws = wb.active; ws.title = 'Confronto Scenari'
    ws.sheet_view.showGridLines = False
    for col,w in [('A',32),('B',18),('C',18),('D',18)]:
        ws.column_dimensions[col].width = w

    def h(row,col,val,color=FGA,bg=BG0,bold=True,size=10,align='left'):
        c=ws.cell(row=row,column=col,value=val)
        c.font=Font(color=color,bold=bold,size=size)
        c.fill=PatternFill('solid',fgColor=bg)
        c.alignment=Alignment(horizontal=align,vertical='center')
        return c
    def row_data(r,label,va,vb,fmt=None,bold=False):
        d = (vb or 0) - (va or 0)
        dc = GRN if d>0 else (RED if d<0 else FGM)
        c1=ws.cell(row=r,column=1,value=label)
        c1.font=Font(color=FGM,size=9,bold=bold); c1.fill=PatternFill('solid',fgColor=BG1 if r%2==0 else BG0)
        for ci,v in [(2,va),(3,vb),(4,d)]:
            c=ws.cell(row=r,column=ci,value=v)
            c.font=Font(color=(PRP if ci==3 else FGA) if ci in (2,3) else dc, bold=bold,size=10)
            c.alignment=Alignment(horizontal='right',vertical='center')
            c.fill=PatternFill('solid',fgColor=BG1 if r%2==0 else BG0)
            if fmt: c.number_format=fmt
        ws.row_dimensions[r].height=16
    def sec(r,label):
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
        c=ws.cell(row=r,column=1,value=f'  {label}')
        c.font=Font(color=FGA,bold=True,size=9); c.fill=PatternFill('solid',fgColor=BG0)
        c.alignment=Alignment(horizontal='left',vertical='center'); ws.row_dimensions[r].height=15

    # Titolo
    ws.merge_cells('A1:D1')
    h(1,1,f'CONFRONTO SCENARI',size=13,align='left')
    ws.row_dimensions[1].height=28
    ws.merge_cells('A2:D2')
    ws.cell(row=2,column=1,value=f'Generato il {datetime.now().strftime("%d/%m/%Y %H:%M")}').font=Font(color=FGM,size=9,italic=True)
    ws.cell(row=2,column=1).fill=PatternFill('solid',fgColor=BG0)
    ws.row_dimensions[2].height=14; ws.row_dimensions[3].height=8

    # Intestazioni colonne
    for ci,(lbl,col) in enumerate([(f'Scenario A — {(sa["created_at"] or "")[:10]}','7c9dff'),
                                    (f'Scenario B — {(sb["created_at"] or "")[:10]}','bc8cff'),
                                    ('Differenza (B − A)','8b949e')],2):
        c=ws.cell(row=4,column=ci,value=lbl)
        c.font=Font(color=col,bold=True,size=9); c.fill=PatternFill('solid',fgColor=BG2)
        c.alignment=Alignment(horizontal='right',vertical='center')
    ws.cell(row=4,column=1,value='Voce').font=Font(color=FGA,bold=True,size=9)
    ws.cell(row=4,column=1).fill=PatternFill('solid',fgColor=BG2)
    ws.row_dimensions[4].height=18

    r=5
    sec(r,'ANAGRAFICA'); r+=1
    for label, ka, kb in [('Offerta','nome_offerta','nome_offerta'),
                           ('Fornitore','nome_fornitore','nome_fornitore'),
                           ('Piano provvigionale','nome_piano','nome_piano'),
                           ('Commodity','commodity','commodity')]:
        c1=ws.cell(row=r,column=1,value=label); c1.font=Font(color=FGM,size=9)
        c1.fill=PatternFill('solid',fgColor=BG1 if r%2==0 else BG0)
        for ci,v in [(2,sa.get(ka,'—')),(3,sb.get(kb,'—'))]:
            c=ws.cell(row=r,column=ci,value=v)
            c.font=Font(color=FG if sa.get(ka)==sb.get(kb) else ('f0883e'),size=10)
            c.alignment=Alignment(horizontal='right',vertical='center')
            c.fill=PatternFill('solid',fgColor=BG1 if r%2==0 else BG0)
        ws.cell(row=r,column=4,value='✓ uguale' if sa.get(ka)==sb.get(kb) else '⚠ diverso').font=Font(color=FGM if sa.get(ka)==sb.get(kb) else 'f0883e',size=9)
        ws.row_dimensions[r].height=16; r+=1

    sec(r,'PARAMETRI TECNICI'); r+=1
    row_data(r,'Spread vendita (€/unità)',   sa['spread_vendita'],     sb['spread_vendita'],     '€#,##0.0000'); r+=1
    row_data(r,'Spread acquisto (€/unità)',  sa['spread_acquisto'],    sb['spread_acquisto'],    '€#,##0.0000'); r+=1
    row_data(r,'Quota fissa vendita (€/mese)',sa['quota_fissa'],       sb['quota_fissa'],        '€#,##0.00'); r+=1
    row_data(r,'Costo gestione PDP (€/mese)',sa['costo_gestione_pdp'],sb['costo_gestione_pdp'], '€#,##0.00'); r+=1
    row_data(r,'Consumo medio',              sa['consumo_medio'],      sb['consumo_medio'],      '#,##0.0000'); r+=1

    sec(r,'COMPOSIZIONE MARGINI'); r+=1
    row_data(r,'Margine spread annuo',  sa['margine_spread_annuo'],sb['margine_spread_annuo'],'€#,##0.00'); r+=1
    row_data(r,'Margine QF annuo',      sa['margine_qf_annuo'],    sb['margine_qf_annuo'],    '€#,##0.00'); r+=1
    row_data(r,'Margine lordo',         sa['margine_lordo'],       sb['margine_lordo'],       '€#,##0.00',bold=True); r+=1

    sec(r,'PROVVIGIONI'); r+=1
    row_data(r,'Provv. agente',         sa['provvigione_agente'],        sb['provvigione_agente'],        '€#,##0.00'); r+=1
    row_data(r,'Provv. sub-agente',     sa['provvigione_sub_agente'],    sb['provvigione_sub_agente'],    '€#,##0.00'); r+=1
    row_data(r,'Provv. area manager',   sa['provvigione_area_manager'],  sb['provvigione_area_manager'],  '€#,##0.00'); r+=1
    row_data(r,'Totale provvigioni',    sa['totale_provvigioni'],        sb['totale_provvigioni'],        '€#,##0.00',bold=True); r+=1

    sec(r,'RISULTATO'); r+=1
    row_data(r,'Margine netto',         sa['margine_netto'],        sb['margine_netto'],       '€#,##0.00',bold=True); r+=1
    row_data(r,'Margine %',             sa['margine_percentuale'],  sb['margine_percentuale'],  '#,##0.0"%"',bold=True); r+=1

    ws.freeze_panes='B5'
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'confronto_{sid_a}_vs_{sid_b}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/admin/prezzi-mercato', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_prezzi_mercato():
    if request.method == 'POST':
        action = request.form.get('action', 'add')
        if action == 'add':
            commodity = request.form.get('commodity', '').upper()
            anno  = int(request.form.get('anno', 0) or 0)
            mese  = int(request.form.get('mese', 0) or 0)
            prezzo = float(request.form.get('prezzo_mwh', 0) or 0)
            fonte  = request.form.get('fonte', 'Manuale').strip() or 'Manuale'
            if commodity in ('LUCE', 'GAS') and 1 <= mese <= 12 and anno >= 2020 and prezzo > 0:
                with get_db() as db:
                    db.execute(
                        'INSERT INTO prezzi_mercato (commodity,anno,mese,prezzo_mwh,fonte,aggiornato_il) '
                        'VALUES (?,?,?,?,?,to_char(NOW(),\'YYYY-MM-DD HH24:MI:SS\')) '
                        'ON CONFLICT (commodity,anno,mese) DO UPDATE SET '
                        'prezzo_mwh=EXCLUDED.prezzo_mwh, fonte=EXCLUDED.fonte, '
                        'aggiornato_il=EXCLUDED.aggiornato_il',
                        (commodity, anno, mese, prezzo, fonte)
                    )
                    db.commit()
                flash(f'Prezzo {"PUN" if commodity=="LUCE" else "PSV"} '
                      f'{MESI_IT[mese]} {anno} aggiornato: {prezzo} €/MWh', 'success')
            else:
                flash('Dati non validi.', 'danger')
        return redirect(url_for('admin_prezzi_mercato'))

    with get_db() as db:
        prezzi_luce = db.execute(
            'SELECT * FROM prezzi_mercato WHERE commodity=? ORDER BY anno DESC,mese DESC LIMIT 24',
            ('LUCE',)
        ).fetchall()
        prezzi_gas = db.execute(
            'SELECT * FROM prezzi_mercato WHERE commodity=? ORDER BY anno DESC,mese DESC LIMIT 24',
            ('GAS',)
        ).fetchall()

    now = datetime.now()
    return render_template('admin_prezzi_mercato.html',
        prezzi_luce=[dict(r) for r in prezzi_luce],
        prezzi_gas=[dict(r) for r in prezzi_gas],
        anno_corrente=now.year,
        mese_corrente=now.month,
        mesi_it=MESI_IT)


@app.route('/admin/prezzi-mercato/elimina/<int:pid>', methods=['POST'])
@login_required
@admin_required
def admin_prezzi_mercato_elimina(pid):
    with get_db() as db:
        db.execute('DELETE FROM prezzi_mercato WHERE id=?', (pid,))
        db.commit()
    flash('Prezzo eliminato.', 'success')
    return redirect(url_for('admin_prezzi_mercato'))


@app.route('/export/confronto-pdf')
@login_required
def export_confronto_pdf():
    sid_a = request.args.get('a', type=int)
    sid_b = request.args.get('b', type=int)
    if not sid_a or not sid_b:
        abort(400)
    anda, andp = uid_and()
    with get_db() as db:
        sa = db.execute(f'SELECT * FROM simulazioni WHERE id=? {anda}', [sid_a]+andp).fetchone()
        sb = db.execute(f'SELECT * FROM simulazioni WHERE id=? {anda}', [sid_b]+andp).fetchone()
    if not sa or not sb:
        abort(404)
    sa, sb = dict(sa), dict(sb)

    # Colori UFLOW
    COLOR_GREEN = colors.HexColor('#10b981')
    COLOR_LIGHT_BG = colors.HexColor('#f0fdf4')

    # Crea PDF
    pdf_file = io.BytesIO()
    doc = SimpleDocTemplate(pdf_file, pagesize=A4, topMargin=0.8*cm, bottomMargin=1.2*cm, leftMargin=1.2*cm, rightMargin=1.2*cm)
    story = []

    # Logo al centro in alto
    try:
        logo_path = os.path.join(os.path.dirname(__file__), 'static/img/uflow-logo.svg')
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=3*cm, height=1.5*cm)
            logo_table = Table([[logo]], colWidths=[17.5*cm])
            logo_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0)
            ]))
            story.append(logo_table)
            story.append(Spacer(1, 0.4*cm))
    except:
        pass

    story.append(Spacer(1, 0.2*cm))

    # Tabella con solo i dati tecnici - usando Paragraph per formattazione corretta
    text_style = ParagraphStyle('normal', fontName='Helvetica', fontSize=8, alignment=0)
    header_style = ParagraphStyle('header', fontName='Helvetica-Bold', fontSize=8, alignment=1)

    data = [
        [Paragraph('VOCE', header_style), Paragraph('SCENARIO A', header_style), Paragraph('SCENARIO B', header_style), Paragraph('DIFFERENZA', header_style)],
        [Paragraph('Offerta', text_style), Paragraph(str(sa.get('nome_offerta', '—')), text_style), Paragraph(str(sb.get('nome_offerta', '—')), text_style), Paragraph('', text_style)],
        [Paragraph('Fornitore', text_style), Paragraph(str(sa.get('nome_fornitore', '—')), text_style), Paragraph(str(sb.get('nome_fornitore', '—')), text_style), Paragraph('', text_style)],
        [Paragraph('Piano', text_style), Paragraph(str(sa.get('nome_piano', '—')), text_style), Paragraph(str(sb.get('nome_piano', '—')), text_style), Paragraph('', text_style)],
        [Paragraph('Spread vendita (€/unità)', text_style), Paragraph(f"{sa.get('spread_vendita', 0):.4f}", text_style), Paragraph(f"{sb.get('spread_vendita', 0):.4f}", text_style), Paragraph(f"{sb.get('spread_vendita', 0) - sa.get('spread_vendita', 0):+.4f}", text_style)],
        [Paragraph('Spread acquisto (€/unità)', text_style), Paragraph(f"{sa.get('spread_acquisto', 0):.4f}", text_style), Paragraph(f"{sb.get('spread_acquisto', 0):.4f}", text_style), Paragraph(f"{sb.get('spread_acquisto', 0) - sa.get('spread_acquisto', 0):+.4f}", text_style)],
        [Paragraph('Quota fissa (€/mese)', text_style), Paragraph(f"{sa.get('quota_fissa', 0):.2f}", text_style), Paragraph(f"{sb.get('quota_fissa', 0):.2f}", text_style), Paragraph(f"{sb.get('quota_fissa', 0) - sa.get('quota_fissa', 0):+.2f}", text_style)],
        [Paragraph('Consumo medio', text_style), Paragraph(f"{sa.get('consumo_medio', 0):.2f}", text_style), Paragraph(f"{sb.get('consumo_medio', 0):.2f}", text_style), Paragraph(f"{sb.get('consumo_medio', 0) - sa.get('consumo_medio', 0):+.2f}", text_style)],
    ]

    table = Table(data, colWidths=[4.5*cm, 3.5*cm, 3.5*cm, 3.5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), COLOR_GREEN),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1fae5')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, COLOR_LIGHT_BG]),
    ]))
    story.append(table)

    # Build PDF
    doc.build(story)
    pdf_file.seek(0)
    return send_file(pdf_file, as_attachment=True,
                     download_name=f'confronto_{sid_a}_vs_{sid_b}.pdf',
                     mimetype='application/pdf')


@app.route('/confronto')
@login_required
def confronto():
    aw, ap = uid_where()
    with get_db() as db:
        sims = db.execute(
            f'SELECT id,nome_offerta,nome_fornitore,nome_piano,commodity,created_at,margine_netto'
            f' FROM simulazioni {aw} ORDER BY created_at DESC LIMIT 100', ap
        ).fetchall()
    sid_a = request.args.get('a', type=int)
    sid_b = request.args.get('b', type=int)
    sim_a = sim_b = None
    costo_a_mensile = costo_b_mensile = []
    costo_a_annuo = costo_b_annuo = 0
    if sid_a and sid_b:
        anda, andp = uid_and()
        with get_db() as db:
            sim_a = db.execute(f'SELECT * FROM simulazioni WHERE id=? {anda}', [sid_a]+andp).fetchone()
            sim_b = db.execute(f'SELECT * FROM simulazioni WHERE id=? {anda}', [sid_b]+andp).fetchone()
        if sim_a: sim_a = dict(sim_a)
        if sim_b: sim_b = dict(sim_b)
        if sim_a and sim_b:
            comm_a = sim_a.get('commodity', 'LUCE')
            prezzi_a = get_prezzi_ultimi_12_mesi(comm_a)
            costo_a_mensile, costo_a_annuo, _, _ = calcola_costo_cliente_12m(sim_a, prezzi_a)
            costo_b_mensile, costo_b_annuo, _, _ = calcola_costo_cliente_12m(sim_b, prezzi_a)
    return render_template('confronto.html',
        sims=[dict(s) for s in sims],
        sim_a=sim_a, sim_b=sim_b,
        sid_a=sid_a, sid_b=sid_b,
        costo_a_mensile=costo_a_mensile, costo_b_mensile=costo_b_mensile,
        costo_a_annuo=costo_a_annuo,     costo_b_annuo=costo_b_annuo,
        mesi_it=MESI_IT)


# ═══════════════════════════════════════════════════════════════════════════════
# ── RETE AGENTI ────────────────────────────────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════════

def _agenti_kpi(aw, ap):
    """Query principale agenti con KPI aggregati (margine, clienti, portafogli)."""
    aw_a = aw.replace('user_id', 'a.user_id') if aw else ''
    with get_db() as db:
        rows = db.execute(f'''
            SELECT a.*,
                pp.nome_piano,
                (par.nome || ' ' || par.cognome) AS nome_superiore,
                par.ruolo AS ruolo_superiore,
                COUNT(DISTINCT pf.id)  AS n_portafogli,
                COUNT(DISTINCT cp.id)  AS n_clienti,
                COALESCE(SUM(cp.margine_netto), 0)  AS tot_margine,
                COALESCE(SUM(cp.margine_lordo), 0)  AS tot_lordo,
                SUM(CASE WHEN cp.margine_netto < 0 THEN 1 ELSE 0 END) AS n_negativi
            FROM agenti a
            LEFT JOIN piani_provvigionali pp ON pp.id = a.piano_id
            LEFT JOIN agenti par ON par.id = a.parent_id
            LEFT JOIN portafogli pf ON pf.agente_id = a.id
            LEFT JOIN clienti_portafoglio cp ON cp.portafoglio_id = pf.id
            {aw_a}
            GROUP BY a.id
            ORDER BY
                CASE a.ruolo WHEN 'DIRETTORE_COMMERCIALE' THEN 1 WHEN 'AREA_MANAGER' THEN 2 WHEN 'AGENTE' THEN 3 ELSE 4 END,
                a.cognome, a.nome
        ''', ap).fetchall()
    out = []
    for r in rows:
        d = dict(r)
        d['pct_margine'] = round(d['tot_margine'] / d['target_margine_annuo'] * 100) \
            if d.get('target_margine_annuo') else None
        d['pct_clienti'] = round(d['n_clienti'] / d['target_clienti'] * 100) \
            if d.get('target_clienti') else None
        out.append(d)
    return out


@app.route('/agenti')
@login_required
def agenti_view():
    aw, ap = uid_where()
    agenti_data = _agenti_kpi(aw, ap)

    with get_db() as db:
        piani = db.execute(
            f'SELECT id, nome_piano FROM piani_provvigionali {aw} ORDER BY nome_piano', ap
        ).fetchall() if aw else db.execute(
            'SELECT id, nome_piano FROM piani_provvigionali ORDER BY nome_piano'
        ).fetchall()
        portafogli_tutti = db.execute(
            f'SELECT id, nome, agente_id FROM portafogli {aw} ORDER BY nome', ap
        ).fetchall() if aw else db.execute(
            'SELECT id, nome, agente_id FROM portafogli ORDER BY nome'
        ).fetchall()

    # KPI globali rete
    n_agenti   = len(agenti_data)
    tot_clienti = sum(a['n_clienti'] for a in agenti_data)
    tot_margine = sum(a['tot_margine'] for a in agenti_data)
    n_dc        = sum(1 for a in agenti_data if a['ruolo'] == 'DIRETTORE_COMMERCIALE')
    n_am        = sum(1 for a in agenti_data if a['ruolo'] == 'AREA_MANAGER')
    n_agt       = sum(1 for a in agenti_data if a['ruolo'] == 'AGENTE')
    n_sub       = sum(1 for a in agenti_data if a['ruolo'] == 'SUB_AGENTE')

    return render_template('agenti.html',
        agenti=agenti_data,
        piani=[dict(p) for p in piani],
        portafogli_tutti=[dict(p) for p in portafogli_tutti],
        n_agenti=n_agenti, n_dc=n_dc, n_am=n_am, n_agt=n_agt, n_sub=n_sub,
        tot_clienti=tot_clienti, tot_margine=tot_margine)


@app.route('/agenti/crea', methods=['POST'])
@login_required
def agente_crea():
    f = request.form
    parent_id = int(f['parent_id']) if f.get('parent_id') else None
    piano_id  = int(f['piano_id'])  if f.get('piano_id')  else None
    with get_db() as db:
        db.execute('''INSERT INTO agenti
            (user_id, nome, cognome, email, telefono, ruolo, zona,
             parent_id, piano_id, data_attivazione, stato, note,
             target_margine_annuo, target_clienti)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (current_user.id,
             f.get('nome','').strip(), f.get('cognome','').strip(),
             f.get('email','').strip(), f.get('telefono','').strip(),
             f.get('ruolo','AGENTE'), f.get('zona','').strip(),
             parent_id, piano_id,
             f.get('data_attivazione',''), f.get('stato','ATTIVO'),
             f.get('note','').strip(),
             float(f.get('target_margine_annuo') or 0),
             int(f.get('target_clienti') or 0)))
        db.commit()
    flash('Agente aggiunto correttamente.', 'success')
    return redirect(url_for('agenti_view'))


@app.route('/agenti/<int:ag_id>/modifica', methods=['POST'])
@login_required
def agente_modifica(ag_id):
    anda, andp = uid_and()
    f = request.form
    parent_id = int(f['parent_id']) if f.get('parent_id') else None
    piano_id  = int(f['piano_id'])  if f.get('piano_id')  else None
    with get_db() as db:
        db.execute(f'''UPDATE agenti SET
            nome=?, cognome=?, email=?, telefono=?, ruolo=?, zona=?,
            parent_id=?, piano_id=?, data_attivazione=?, stato=?,
            note=?, target_margine_annuo=?, target_clienti=?
            WHERE id=? {anda}''',
            [f.get('nome','').strip(), f.get('cognome','').strip(),
             f.get('email','').strip(), f.get('telefono','').strip(),
             f.get('ruolo','AGENTE'), f.get('zona','').strip(),
             parent_id, piano_id,
             f.get('data_attivazione',''), f.get('stato','ATTIVO'),
             f.get('note','').strip(),
             float(f.get('target_margine_annuo') or 0),
             int(f.get('target_clienti') or 0),
             ag_id] + andp)
        db.commit()
    flash('Agente aggiornato.', 'success')
    return redirect(url_for('agenti_view'))


@app.route('/agenti/<int:ag_id>/elimina', methods=['POST'])
@login_required
def agente_elimina(ag_id):
    anda, andp = uid_and()
    with get_db() as db:
        db.execute('UPDATE portafogli SET agente_id=NULL WHERE agente_id=?', (ag_id,))
        db.execute('UPDATE agenti SET parent_id=NULL WHERE parent_id=?', (ag_id,))
        db.execute(f'DELETE FROM agenti WHERE id=? {anda}', [ag_id] + andp)
        db.commit()
    flash('Agente rimosso dalla rete.', 'success')
    return redirect(url_for('agenti_view'))


@app.route('/agenti/<int:ag_id>/assegna-portafoglio', methods=['POST'])
@login_required
def agente_assegna_portafoglio(ag_id):
    anda, andp = uid_and()
    pf_ids = request.form.getlist('portafoglio_ids')
    with get_db() as db:
        # Prima rimuovi tutti i portafogli già assegnati a questo agente
        db.execute('UPDATE portafogli SET agente_id=NULL WHERE agente_id=?', (ag_id,))
        # Poi assegna quelli selezionati
        for pf_id in pf_ids:
            db.execute(f'UPDATE portafogli SET agente_id=? WHERE id=? {anda}',
                       [ag_id, int(pf_id)] + andp)
        db.commit()
    flash('Portafogli aggiornati.', 'success')
    return redirect(url_for('agenti_view'))


@app.route('/api/agenti/tree')
@login_required
def api_agenti_tree():
    aw, ap = uid_where()
    aw_a = aw.replace('user_id', 'a.user_id') if aw else ''
    with get_db() as db:
        rows = db.execute(f'''
            SELECT a.id, a.nome, a.cognome, a.ruolo, a.parent_id, a.stato, a.zona,
                COALESCE(SUM(cp.margine_netto), 0) AS tot_margine,
                COUNT(DISTINCT cp.id)  AS n_clienti,
                COUNT(DISTINCT pf.id)  AS n_portafogli,
                a.target_margine_annuo, a.target_clienti
            FROM agenti a
            LEFT JOIN portafogli pf ON pf.agente_id = a.id
            LEFT JOIN clienti_portafoglio cp ON cp.portafoglio_id = pf.id
            {aw_a}
            GROUP BY a.id
        ''', ap).fetchall()

    nodes = {}
    for r in rows:
        d = dict(r)
        d['children'] = []
        d['pct_margine'] = round(d['tot_margine'] / d['target_margine_annuo'] * 100) \
            if d.get('target_margine_annuo') else None
        nodes[d['id']] = d

    roots = []
    for node in nodes.values():
        pid = node.get('parent_id')
        if pid and pid in nodes:
            nodes[pid]['children'].append(node)
        else:
            roots.append(node)

    # Sort within each level
    def sort_nodes(nl):
        nl.sort(key=lambda x: (x['cognome'] or '', x['nome'] or ''))
        for n in nl:
            sort_nodes(n['children'])
    sort_nodes(roots)

    return jsonify(roots)


if __name__ == '__main__':
    print('\n✅ Energia Simulator avviato!')
    print('📌 Apri il browser su: http://127.0.0.1:5000\n')
    app.run(debug=True, port=5000)
